# -*- coding: utf-8 -*-
from .base import *  # ctk, tk, ttk, messagebox, filedialog, BASE_DIR
import os, json, csv, sys, re
from datetime import datetime
from .machine_recipes_constants import *
from .machine_recipes_constants import (
    _load_json,
    _save_json,
    _save_version_snapshot,
)
from .machine_recipes_history_panel import open_history

class MachineRecipesView(ctk.CTkFrame):
    """Panel de recetas de máquina con layout SELOGICA + historial/versiones."""

    def __init__(self, master, app):
        super().__init__(master, fg_color="transparent")
        self.app = app
        self.vars = {}          # id -> StringVar
        self.current_mold = ""
        self._dirty = False
        self._preview_version = None
        self._status_var = tk.StringVar(value="")
        self._motivo_ph = (
            "Formato sugerido: [Qué cambiaste] -> [Por qué] -> [Resultado esperado].\n"
            "Ej.: Aumenté HP de 60->70 bar para mejorar compacción; espero eliminar porosidad en esquina A."
        )
        self._build()

    # --------------------------- UI ---------------------------
    def _build(self):
        # Header
        header = ctk.CTkFrame(self, corner_radius=0, fg_color=("white", "#0e1117"))
        header.pack(fill="x", side="top")
        left = ctk.CTkFrame(header, fg_color="transparent"); left.pack(side="left", padx=16, pady=10)
        ctk.CTkButton(left, text="← Menú", width=110, corner_radius=10,
                      fg_color="#E5E7EB", text_color="#111", hover_color="#D1D5DB",
                      command=self._back_menu_guard).pack(side="left", padx=(0, 10))
        title = ctk.CTkFrame(left, fg_color="transparent"); title.pack(side="left")
        ctk.CTkLabel(title, text="Recetas de Máquina", font=ctk.CTkFont("Helvetica", 22, "bold")).pack(anchor="w")
        ctk.CTkLabel(title, text="Layout estilo SELOGICA • Historial y versiones",
                     text_color=("#6b7280", "#9CA3AF")).pack(anchor="w")

        # Toolbar
        tools = ctk.CTkFrame(self, fg_color="transparent"); tools.pack(fill="x", padx=16, pady=(8, 0))
        opciones = ["— Selecciona —"] + list(getattr(self.app, "recipe_map", {}).keys())
        self.mold_var = tk.StringVar(value=opciones[0])
        self.mold_menu = ctk.CTkOptionMenu(tools, values=opciones, variable=self.mold_var, width=260,
                                           command=lambda *_: self._on_pick_mold())
        self.mold_menu.pack(side="left", padx=(0, 8))
        ctk.CTkButton(tools, text="📥 Importar Excel", command=self._import_excel).pack(side="left", padx=6)
        ctk.CTkButton(tools, text="🕓 Historial / Versiones", command=self._open_history).pack(side="left", padx=6)
        ctk.CTkButton(tools, text="📂 Carpeta de recetas",
                      command=lambda: self._open_folder(RECIPES_DIR)).pack(side="left", padx=6)

        # === DOCUMENTAR CAMBIO ===
        doc_card = ctk.CTkFrame(self, corner_radius=12, fg_color=("white", "#111827"))
        doc_card.pack(fill="x", padx=16, pady=(10, 0))
        topbar = ctk.CTkFrame(doc_card, fg_color="transparent"); topbar.pack(fill="x", padx=12, pady=(10, 4))
        ctk.CTkLabel(topbar, text="Documentar cambio", font=ctk.CTkFont(size=14, weight="bold")).pack(side="left")
        actions = ctk.CTkFrame(topbar, fg_color="transparent"); actions.pack(side="right")
        ctk.CTkButton(actions, text="💾 Guardar (Ctrl+S)", command=self._save).pack(side="left")
        ctk.CTkButton(actions, text="🗑", fg_color="#ef4444", hover_color="#dc2626", width=44,
                      command=self._clear_all).pack(side="left", padx=(8,0))
        body = ctk.CTkFrame(doc_card, fg_color="transparent"); body.pack(fill="both", padx=12, pady=(0, 12))
        self.motivo_txt = ctk.CTkTextbox(body, height=110)
        self.motivo_txt.pack(fill="x", expand=False)
        self._init_motivo_placeholder()
        btns = ctk.CTkFrame(body, fg_color="transparent"); btns.pack(fill="x", pady=(6,0))
        ctk.CTkButton(btns, text="Ejemplo", width=80, command=self._insert_motivo_example).pack(side="left")

        # status
        ctk.CTkLabel(self, textvariable=self._status_var,
                     text_color=("#6b7280", "#9CA3AF"), anchor="w").pack(fill="x", padx=16, pady=(4,0))

        # CONTENIDO con scroll (formulario)
        self.scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.scroll.pack(fill="both", expand=True, padx=16, pady=16)

        sheet = ctk.CTkFrame(self.scroll, corner_radius=14, fg_color=("white", "#111827"))
        sheet.pack(fill="x", padx=2, pady=2)
        sheet.grid_columnconfigure(0, weight=3)
        sheet.grid_columnconfigure(1, weight=2)

        hdr = self._card(sheet, "Parameter overview (Cabecera)", row=0, col=0)
        self._header_grid(hdr)
        key = self._card(sheet, "Key data", row=0, col=1)
        self._keydata_table(key)
        blocks = self._card(sheet, "", row=1, col=0, col_span=2)
        self._full_left_side(blocks)

        self._register_shortcuts_once()

    # ----- small UI builders -----
    def _card(self, parent, title, row=0, col=0, col_span=1):
        card = ctk.CTkFrame(parent, corner_radius=14, fg_color=("white", "#111827"))
        card.grid(row=row, column=col, columnspan=col_span, padx=8, pady=8, sticky="nsew")
        if title:
            ctk.CTkLabel(card, text=title, font=ctk.CTkFont("Helvetica", 13, "bold")).pack(anchor="w", padx=10, pady=(10, 0))
            ctk.CTkFrame(card, height=1, fg_color=("#E5E7EB", "#2B2B2B")).pack(fill="x", padx=10, pady=(6, 10))
        return card

    def _lbl(self, parent, text, **grid):
        l = ctk.CTkLabel(parent, text=text); l.grid(**grid); return l

    def _ent(self, parent, fid, w=110, justify="center", **grid):
        v = self.vars.setdefault(fid, tk.StringVar())
        e = ctk.CTkEntry(parent, textvariable=v, width=w, justify=justify)
        e.grid(**grid)
        v.trace_add("write", lambda *_: self._mark_dirty())
        return e

    def _header_grid(self, parent):
        g = ctk.CTkFrame(parent, fg_color="transparent"); g.pack(fill="x", padx=10, pady=10)
        for i in range(6): g.grid_columnconfigure(i, weight=1)
        self._lbl(g, "Program",        row=0, column=0, sticky="w", padx=4, pady=4)
        self._ent(g, "program",        row=0, column=1, sticky="ew", padx=4, pady=4)
        self._lbl(g, "Date of entry:", row=0, column=2, sticky="w", padx=4, pady=4)
        self._ent(g, "date_of_entry",  row=0, column=3, sticky="ew", padx=4, pady=4)
        self._lbl(g, "Cavities",       row=0, column=4, sticky="w", padx=4, pady=4)
        self._ent(g, "cavities",       row=0, column=5, sticky="ew", padx=4, pady=4)
        self._lbl(g, "Mould desig.",   row=1, column=0, sticky="w", padx=4, pady=4)
        self._ent(g, "mould_desig",    row=1, column=1, sticky="ew", padx=4, pady=4)
        self._lbl(g, "Machine",        row=1, column=2, sticky="w", padx=4, pady=4)
        self._ent(g, "machine",        row=1, column=3, sticky="ew", padx=4, pady=4)
        self._lbl(g, "Material",       row=2, column=0, sticky="w", padx=4, pady=4)
        self._ent(g, "material",       row=2, column=1, sticky="ew", padx=4, pady=4)

    def _keydata_table(self, parent):
        grid = ctk.CTkFrame(parent, fg_color="transparent"); grid.pack(fill="x", padx=10, pady=10)
        for i in range(2): grid.grid_columnconfigure(i, weight=1)
        rows = [
            ("cycle_time_s",           "Cycle time [s]"),
            ("injection_time_s",       "Injection time [s]"),
            ("holding_press_time_s",   "Holding press. time [s]"),
            ("rem_cooling_time_s",     "Rem. cooling time [s]"),
            ("dosage_time_s",          "Dosage time [s]"),
            ("screw_stroke_mm",        "Screw stroke [mm]"),
            ("mould_stroke_mm",        "Mould stroke [mm]"),
            ("ejector_stroke_mm",      "Ejector stroke [mm]"),
            ("shot_weight_g",          "Shot weight [g]"),
            ("plasticising_flow_kgh",  "Plasticising flow [kg/h]"),
            ("dosage_capacity_gs",     "Dosage capacity [g/s]"),
            ("dosage_volume_ccm",      "Dosage volume [ccm]"),
            ("material_cushion_ccm",   "Material cushion [ccm]"),
            ("max_inj_pressure_bar",   "max. inj. pressure [bar]"),
        ]
        for r, (fid, label) in enumerate(rows):
            self._lbl(grid, label, row=r, column=0, sticky="w", padx=4, pady=4)
            self._ent(grid, fid, row=r, column=1, sticky="ew", padx=4, pady=4)

    def _full_left_side(self, parent):
        ctk.CTkLabel(parent, text="Injection unit", font=ctk.CTkFont("Helvetica", 12, "bold")
                     ).pack(anchor="center", pady=(4, 0))
        inj_unit = ctk.CTkFrame(parent, fg_color="transparent"); inj_unit.pack(fill="x", padx=10, pady=(0, 8))
        inj_unit.grid_columnconfigure((0,1,2,3,4), weight=1)
        self._lbl(inj_unit, "Screw Ø [mm]", row=0, column=0, sticky="e", padx=6, pady=4)
        self._ent(inj_unit, "screw_d_mm", w=80, row=0, column=1, sticky="w", padx=4, pady=4)
        self._lbl(inj_unit, "Pcs. 1", row=0, column=3, sticky="e", padx=6, pady=4)
        self._ent(inj_unit, "pcs_1", w=80, row=0, column=4, sticky="w", padx=4, pady=4)

        self._section_table(parent, "Injection", [
            ("Injection press. limiting [bar]", ["inj_press_lim_1", "inj_press_lim_2", "inj_press_lim_3"]),
            ("Injection speed [mm/s]",          ["inj_speed_1", "inj_speed_2", "inj_speed_3"]),
            ("End of stage [mm]",               ["inj_end_stage_mm_1", "inj_end_stage_mm_2", "inj_end_stage_mm_3"]),
            ("Injection flow [ccm/s]",          ["inj_flow_1", "inj_flow_2", "inj_flow_3"]),
            ("End of stage [ccm]",              ["inj_end_stage_ccm_1", "inj_end_stage_ccm_2", "inj_end_stage_ccm_3"]),
        ])

        self._section_table(parent, "Plasticizing (St.1)", [
            ("Screw speed [m/min]",             ["plast_screw_speed"]),
            ("Back pressure [bar]",             ["plast_back_pressure"]),
            ("End of stage [ccm]",              ["plast_end_stage_ccm"]),
        ])

        self._section_table(parent, "Holding pressure (Pcs.2)", [
            ("Time [s]",                        ["hp_time_1", "hp_time_2", "hp_time_3"]),
            ("Pressure [bar]",                  ["hp_press_1", "hp_press_2", "hp_press_3", "hp_press_4"]),
        ])

        self._section_table(parent, "Temperatures (1..5)", [
            ("Cylinder temp. [°C]",             ["temp_c1", "temp_c2", "temp_c3", "temp_c4", "temp_c5"]),
            ("Tolerances [°C]",                 ["tol_c1", "tol_c2", "tol_c3", "tol_c4", "tol_c5"]),
            ("Feed yoke temperature [°C]",      ["feed_yoke_temp"]),
            ("Lower enable tol. [°C]",          ["lower_enable_tol"]),
            ("Upper switch-off tol. [°C]",      ["upper_switch_off_tol"]),
        ])

        self._section_table(parent, "Mould movements — Opening (St.1 / St.2 / St.3)", [
            ("End of stage [mm]",               ["open_end_mm_1", "open_end_mm_2", "open_end_mm_3"]),
            ("Speed [mm/s]",                    ["open_speed_1", "open_speed_2", "open_speed_3"]),
            ("Force [kN]",                      ["open_force_1", "open_force_2", "open_force_3"]),
        ])

        self._section_table(parent, "Mould movements — Closing (St.1 / St.2 / St.3 / An. HD)", [
            ("End of stage [mm]",               ["close_end_mm_1", "close_end_mm_2", "close_end_mm_3", "close_end_mm_4"]),
            ("Speed [mm/s]",                    ["close_speed_1", "close_speed_2", "close_speed_3", "close_speed_4"]),
            ("Force [kN]",                      ["close_force_1", "close_force_2", "close_force_3"]),
        ])

        self._section_table(parent, "Clamping", [
            ("Mould closed [kN]",               ["mould_closed_kn"]),
        ])

    def _section_table(self, parent, title, rows):
        card = ctk.CTkFrame(parent, corner_radius=12, fg_color=("white", "#111a27"))
        card.pack(fill="x", padx=10, pady=8)
        ctk.CTkLabel(card, text=title, font=ctk.CTkFont("Helvetica", 12, "bold")).pack(anchor="w", padx=10, pady=(8, 6))
        grid = ctk.CTkFrame(card, fg_color="transparent"); grid.pack(fill="x", padx=10, pady=(0,10))
        max_cols = max(len(cols) for _, cols in rows)
        for i in range(max_cols + 1): grid.grid_columnconfigure(i, weight=1)
        for r, (label, ids) in enumerate(rows):
            self._lbl(grid, label, row=r, column=0, sticky="w", padx=4, pady=4)
            for c, fid in enumerate(ids, start=1):
                self._ent(grid, fid, w=90, row=r, column=c, sticky="ew", padx=4, pady=4)

    # --------------------------- Eventos básicos ---------------------------
    def _register_shortcuts_once(self):
        root = self.winfo_toplevel()
        if getattr(root, "_recipes_shortcuts_ok", False): return
        root.bind("<Control-s>", lambda e: (self._save(), "break"), add="+")
        root._recipes_shortcuts_ok = True

    def _init_motivo_placeholder(self):
        # placeholder gris
        self._motivo_placeholder_on = True
        self.motivo_txt.delete("1.0", "end")
        self.motivo_txt.insert("1.0", self._motivo_ph)
        self.motivo_txt.configure(text_color=("#9CA3AF", "#6b7280"))
        self.motivo_txt.bind("<FocusIn>", self._motivo_focus_in)
        self.motivo_txt.bind("<FocusOut>", self._motivo_focus_out)

    def _motivo_focus_in(self, *_):
        if self._motivo_placeholder_on:
            self.motivo_txt.delete("1.0", "end")
            self.motivo_txt.configure(text_color=("black", "white"))
            self._motivo_placeholder_on = False

    def _motivo_focus_out(self, *_):
        content = (self.motivo_txt.get("1.0", "end").strip())
        if not content:
            self._init_motivo_placeholder()

    def _get_motivo_text(self) -> str:
        if self._motivo_placeholder_on: return ""
        return self.motivo_txt.get("1.0", "end").strip()

    def _set_motivo_text(self, text: str):
        self._motivo_placeholder_on = False
        self.motivo_txt.delete("1.0", "end")
        self.motivo_txt.insert("1.0", text)
        self.motivo_txt.configure(text_color=("black", "white"))

    def _insert_motivo_example(self):
        example = ("Ajusté velocidad de inyección 50->65 mm/s para mejorar llenado en nervaduras; "
                   "espero reducir cortinas y estabilizar peso del disparo (+0.2g).")
        cur = self._get_motivo_text()
        if not cur:
            self._set_motivo_text(example)
        else:
            if not cur.endswith("\n"):
                self.motivo_txt.insert("end", "\n")
            self.motivo_txt.insert("end", example)

    def _mark_dirty(self): self._dirty = True

    def _back_menu_guard(self):
        if self._dirty and not messagebox.askyesno("Cambios sin guardar", "Tienes cambios sin guardar. ¿Descartar?"):
            return
        self._dirty = False
        self._preview_version = None
        self._status_var.set("")
        if hasattr(self.app, "go_menu"):
            self.app.go_menu()

    def _on_pick_mold(self):
        mid = (self.mold_var.get() or "").strip()
        self.current_mold = "" if (not mid or mid.startswith("—")) else mid
        for v in self.vars.values(): v.set("")
        self._dirty = False
        self._preview_version = None
        self._status_var.set("")
        self._init_motivo_placeholder()
        if self.current_mold:
            data = _load_json(self.current_mold)
            for k, v in self.vars.items():
                v.set(str(data.get(k, "")))
            if "date_of_entry" in self.vars and not self.vars["date_of_entry"].get().strip():
                self.vars["date_of_entry"].set(datetime.now().strftime("%Y-%m-%d"))

    def _collect(self) -> dict:
        out = {}
        for k, v in self.vars.items():
            val = (v.get() or "").strip()
            if k in NUM_FIELDS: val = _cast_numeric(val)
            out[k] = val
        return out

    def _resolve_usuario(self) -> str:
        try:
            op = getattr(self.app, "operador", "")
            return op.get() if hasattr(op, "get") else str(op or "")
        except Exception:
            return ""

    def _save(self, *_):
        if not self.current_mold:
            messagebox.showwarning("Molde", "Selecciona un molde."); return
        motivo = self._get_motivo_text()
        if not motivo:
            messagebox.showwarning("Motivo requerido", "Escribe el motivo del cambio para guardar."); return
        if "date_of_entry" in self.vars and not self.vars["date_of_entry"].get().strip():
            self.vars["date_of_entry"].set(datetime.now().strftime("%Y-%m-%d"))

        old = _load_json(self.current_mold)
        new = self._collect()

        diffs = []
        for k, nv in new.items():
            ov = str(old.get(k, ""))
            if (nv or "") != (ov or ""):
                diffs.append(f"{k}: '{ov}' -> '{nv}'")
        if not diffs:
            messagebox.showinfo("Recetas de Máquina", "No hubo cambios en los parámetros.")
            return

        _save_json(self.current_mold, new)

        usuario = self._resolve_usuario()
        diffs_text = " | ".join(diffs)
        ver = _save_version_snapshot(self.current_mold, new, usuario, motivo, diffs_text)

        exists = os.path.exists(HISTORY_CSV)
        with open(HISTORY_CSV, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if not exists:
                w.writerow(["ts", "molde_id", "usuario", "motivo", "cambios", "version"])
            w.writerow([datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        self.current_mold, usuario, motivo, diffs_text, ver])

        self._init_motivo_placeholder()
        self._dirty = False
        self._preview_version = None
        self._status_var.set(f"Guardado como {ver}.")
        messagebox.showinfo("Recetas de Máquina", f"Parámetros guardados.\nSe creó snapshot: {ver}")

    def _clear_all(self):
        if not self.current_mold:
            if messagebox.askyesno("Limpiar", "¿Limpiar todos los campos visibles?"):
                for v in self.vars.values(): v.set("")
            return
        if messagebox.askyesno("Limpiar", f"¿Limpiar todos los campos del molde {self.current_mold}?"):
            for v in self.vars.values(): v.set("")
            self._dirty = True

    # --------------------------- Historial / Versiones ---------------------------
    def _open_history(self):
        open_history(self)

    # --------------------------- Importar Excel simple ---------------------------
    def _import_excel(self):
        try:
            from openpyxl import load_workbook
        except Exception:
            messagebox.showwarning("Falta dependencia", "Instala openpyxl: pip install openpyxl")
            return
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx;*.xlsm;*.xls")])
        if not path: return
        try:
            wb = load_workbook(path, data_only=True, keep_vba=True)
            ws = wb.active
            hits, misses = 0, 0
            for excel_key, spec in EXCEL_MAP.items():
                ui_key = ALIAS_EXCEL_TO_UI.get(excel_key, excel_key)
                try:
                    addr = _anchor_address(ws, spec)
                    val = ws[addr].value
                except Exception:
                    val = None
                if ui_key in self.vars and val is not None:
                    self.vars[ui_key].set(_cast_numeric(val) if ui_key in NUM_FIELDS else str(val))
                    hits += 1
                else:
                    misses += 1
            self._dirty = True
            messagebox.showinfo("Importar",
                                f"Se mapearon {hits} campo(s) desde el Excel.\n"
                                f"Sin valor o no presente en UI: {misses}.")
        except Exception as e:
            messagebox.showerror("Importar Excel", f"No se pudo leer el archivo:\n{e}")

    # --------------------------- Utilidades ---------------------------
    def _open_folder(self, path):
        try:
            if os.path.isdir(path):
                if sys.platform.startswith("win"):
                    os.startfile(path)
                elif sys.platform == "darwin":
                    os.system(f'open "{path}"')
                else:
                    os.system(f'xdg-open "{path}"')
            else:
                folder = os.path.dirname(path)
                if sys.platform.startswith("win"):
                    os.startfile(folder)
                elif sys.platform == "darwin":
                    os.system(f'open "{folder}"')
                else:
                    os.system(f'xdg-open "{folder}"')
        except Exception as e:
            messagebox.showwarning("Abrir carpeta", f"No se pudo abrir:\n{e}")
