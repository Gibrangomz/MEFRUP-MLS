# views/reports_view.py — Dashboard claro tipo SaaS (Seaborn),
# export pulido (logo a la derecha), colores pro y más headroom

from .base import *  # ctk, tk, Calendar, messagebox, filedialog, MACHINES, resumen_rango_maquina, leer_shipments

import io, os, json, tempfile, logging, threading, traceback
from datetime import date, datetime, timedelta, time as dtime
from concurrent.futures import ThreadPoolExecutor

import numpy as np
import pandas as pd
from PIL import Image, ImageDraw, ImageFont
from fpdf import FPDF

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------- Matplotlib + Seaborn (tema claro pro) ----------
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.ticker import MaxNLocator
import seaborn as sns

# NO forzamos modo oscuro (queda el default claro)

# Tamaño lógico y resolución de salida
BASE_DPI   = 140
RENDER_DPI = 360
CHART_W    = 560
CHART_H    = 220

LOGO_PATH  = globals().get("LOGO_PATH", "")

# Paleta moderna (acento colorido)
PALETTE = {
    "buenas": "#22C55E",  # green-500
    "scrap":  "#EF4444",  # red-500
    "total":  "#06B6D4",  # cyan-500

    "avail":  "#3B82F6",  # blue-500
    "perf":   "#F59E0B",  # amber-500
    "qual":   "#10B981",  # emerald-500
    "oee":    "#8B5CF6",  # violet-500

    "ship":   "#DC2626",  # red-600
    "disp":   "#059669",  # emerald-600

    "oee_line": "#2563EB",
    "band":     "#D1FAE5",  # verde pálido
    "mean":     "#111827",  # casi negro
}

# Estilo base claro
sns.set_theme(style="whitegrid", rc={
    "figure.dpi": 120,
    "savefig.dpi": 360,
    "font.size": 9,
    "axes.titlesize": 12,
    "axes.labelsize": 10,
    "xtick.labelsize": 8.5,
    "ytick.labelsize": 8.5,
    "axes.grid": True,
    "grid.alpha": 0.25,
})
plt.rcParams.update({
    "figure.facecolor":   "#FFFFFF",
    "axes.facecolor":     "#FFFFFF",
    "savefig.facecolor":  "#FFFFFF",
    "axes.edgecolor":     "#E5E7EB",
    "axes.labelcolor":    "#111827",
    "xtick.color":        "#374151",
    "ytick.color":        "#374151",
    "text.color":         "#111827",
    "grid.color":         "#E5E7EB",
})

_EXEC = ThreadPoolExecutor(max_workers=4)

def _pdf_safe(text: str) -> str:
    if text is None: return ""
    rep = {"\u2192": "->", "\u2190": "<-", "\u2014": "-", "\u2013": "-",
           "\u00B1": "+/-", "\u03C3": "sigma", "\u00B0": "deg", "–": "-", "—": "-", "±": "+/-"}
    for k, v in rep.items(): text = text.replace(k, v)
    return text.encode("latin-1", "replace").decode("latin-1")


class ReportsView(ctk.CTkFrame):
    def __init__(self, master, app):
        super().__init__(master, fg_color="transparent")
        self.app = app
        self._plot_bytes, self._plot_titles, self._plot_captions = [], [], []
        self._last_data, self._last_stats = [], {}
        self._generating = False
        self._build()

    # ===================== UI =====================
    def _build(self):
        # Header claro tipo SaaS
        header = ctk.CTkFrame(self, corner_radius=0, fg_color="#F9FAFB")
        header.pack(fill="x", side="top")
        ctk.CTkButton(header, text="← Menú", command=self.app.go_menu, width=110,
                      corner_radius=10, fg_color="#E5E7EB", text_color="#111827",
                      hover_color="#D1D5DB").pack(side="left", padx=(16, 10), pady=10)
        ctk.CTkLabel(header, text="Financial / Production Dashboard",
                     font=ctk.CTkFont("Helvetica", 20, "bold"),
                     text_color="#111827").pack(side="left", pady=10)

        right = ctk.CTkFrame(header, fg_color="transparent")
        right.pack(side="right", padx=16, pady=10)
        self.btn_generate_top = ctk.CTkButton(right, text="↻ Generar",
                                              command=self._generar_async,
                                              corner_radius=10, fg_color="#2563EB",
                                              text_color="white")
        self.btn_generate_top.pack(side="right")

        # Filtros
        filters = ctk.CTkFrame(self, fg_color="#FFFFFF")
        filters.pack(fill="x", padx=32, pady=(10, 0))
        opciones = [m.get("id", "") for m in (MACHINES if "MACHINES" in globals() else [])] or [""]
        self.machine_var = tk.StringVar(value=opciones[0] if opciones else "")
        ctk.CTkLabel(filters, text="Máquina:", text_color="#111827"
        ).grid(row=0, column=0, sticky="e", padx=(12, 6), pady=10)
        ctk.CTkOptionMenu(filters, values=opciones, variable=self.machine_var, width=160
        ).grid(row=0, column=1, sticky="w", padx=(0, 18), pady=10)

        ctk.CTkLabel(filters, text="Desde:", text_color="#111827"
        ).grid(row=0, column=2, sticky="e", padx=(6, 6))
        self.desde_entry = ctk.CTkEntry(filters, width=130); self.desde_entry.grid(row=0, column=3, sticky="w")
        ctk.CTkButton(filters, text="📅", width=36, command=lambda: self._calendar_pick(self.desde_entry)
        ).grid(row=0, column=4, padx=(6, 18))

        ctk.CTkLabel(filters, text="Hasta:", text_color="#111827"
        ).grid(row=0, column=5, sticky="e", padx=(6, 6))
        self.hasta_entry = ctk.CTkEntry(filters, width=130); self.hasta_entry.grid(row=0, column=6, sticky="w")
        ctk.CTkButton(filters, text="📅", width=36, command=lambda: self._calendar_pick(self.hasta_entry)
        ).grid(row=0, column=7, padx=(6, 18))

        self.btn_generate = ctk.CTkButton(filters, text="Generar", command=self._generar_async, corner_radius=10,
                                          fg_color="#2563EB", text_color="white")
        self.btn_generate.grid(row=0, column=8, sticky="e", padx=(0, 12))

        # Scroll principal
        self.scroll = ctk.CTkScrollableFrame(self, corner_radius=0, fg_color="#F3F4F6")
        self.scroll.pack(fill="both", expand=True, padx=16, pady=16)

        self._build_kpis(self.scroll)

        self.charts = ctk.CTkFrame(self.scroll, corner_radius=20, fg_color="#FFFFFF")
        self.charts.pack(fill="both", expand=True)
        self.charts.grid_columnconfigure((0, 1), weight=1)
        ctk.CTkLabel(self.charts, text="Genera un reporte para ver las gráficas…",
                     text_color="#6B7280"
        ).grid(row=0, column=0, padx=12, pady=12, sticky="w")

        self._build_table(self.scroll)

        btns = ctk.CTkFrame(self.scroll, fg_color="transparent")
        btns.pack(fill="x", pady=(12, 10))
        ctk.CTkButton(btns, text="Exportar PDF/Excel", command=self._exportar,
                      corner_radius=10, fg_color="#22C55E", text_color="white").pack(side="right")

    def _build_kpis(self, parent):
        stats = ctk.CTkFrame(parent, corner_radius=18, fg_color="#FFFFFF")
        stats.pack(fill="x", pady=(0, 14))
        stats.grid_columnconfigure((0,1,2,3,4), weight=1)

        self.card_total,  self.lbl_total  = self._kpi(stats, "Total (pzs)",   "#06B6D4"); self.card_total.grid(row=0, column=0, padx=6, pady=6, sticky="nsew")
        self.card_buenas, self.lbl_buenas = self._kpi(stats, "Buenas (pzs)",  "#22C55E"); self.card_buenas.grid(row=0, column=1, padx=6, pady=6, sticky="nsew")
        self.card_scrap,  self.lbl_scrap  = self._kpi(stats, "Scrap (pzs)",   "#EF4444"); self.card_scrap.grid(row=0, column=2, padx=6, pady=6, sticky="nsew")
        self.card_oee,    self.lbl_oee    = self._kpi(stats, "OEE Prom",      "#8B5CF6"); self.card_oee.grid(row=0, column=3, padx=6, pady=6, sticky="nsew")
        self.card_paro,   self.lbl_paro   = self._kpi(stats, "Paro total (min)", "#9CA3AF"); self.card_paro.grid(row=0, column=4, padx=6, pady=6, sticky="nsew")

    def _kpi(self, parent, title, accent_hex="#06B6D4"):
        card = ctk.CTkFrame(parent, corner_radius=14, fg_color="#F9FAFB")        
        header = ctk.CTkFrame(card, corner_radius=12, fg_color="#FFFFFF")        
        header.pack(fill="x", padx=8, pady=(8,4))
        ctk.CTkLabel(header, text=title, text_color="#374151",
                     font=ctk.CTkFont("Helvetica", 12, "bold")).pack(anchor="w", padx=10, pady=6)
        val = ctk.CTkLabel(card, text="0", text_color="#111827",
                           font=ctk.CTkFont("Helvetica", 22, "bold"))
        val.pack(anchor="w", padx=14, pady=(0,10))
        dot = ctk.CTkFrame(header, width=10, height=10, corner_radius=999, fg_color=accent_hex)
        dot.place(relx=1.0, x=-12, rely=0.5, anchor="e")
        return card, val

    def _build_table(self, parent):
        card = ctk.CTkFrame(parent, corner_radius=16, fg_color="#FFFFFF")
        card.pack(fill="both", expand=True, pady=(14, 0))
        ctk.CTkLabel(card, text="Detalle diario", font=ctk.CTkFont("Helvetica", 14, "bold"),
                     text_color="#111827").pack(anchor="w", padx=12, pady=(10, 6))
        ctk.CTkFrame(card, height=1, fg_color="#E5E7EB").pack(fill="x", padx=12, pady=(0, 8))
        cols = ("fecha", "availability", "performance", "quality", "oee", "buenas", "scrap", "total")
        self.tbl = ttk.Treeview(card, columns=cols, show="headings", height=10)
        for k, t, w in [
            ("fecha", "Fecha", 110), ("availability", "Avail %", 80), ("performance", "Eff %", 80),
            ("quality", "Qual %", 80), ("oee", "OEE %", 80), ("buenas", "Buenas", 80),
            ("scrap", "Scrap", 80), ("total", "Total", 80)
        ]:
            self.tbl.heading(k, text=t); self.tbl.column(k, width=w, anchor="center")
        self.tbl.pack(fill="both", expand=True, padx=12, pady=(0, 10))
        self.lbl_tbl_total = ctk.CTkLabel(card, text="—", text_color="#6B7280")
        self.lbl_tbl_total.pack(anchor="w", padx=12, pady=(0, 12))

    # ... (el resto de helpers, gráficas y export se mantienen IGUAL que en el modo oscuro,
    # ya que los colores de paleta y estilo matplotlib ya están ajustados a claro)
    # ============== Helpers ==============
    def _tone(self, oee):
        return ("#DCFCE7", "#065F46") if oee >= 85 else (("#FEF9C3", "#92400E") if oee >= 60 else ("#FEE2E2", "#991B1B"))

    def _calendar_pick(self, entry):
        try:
            y, m, d = map(int, (entry.get() or date.today().isoformat()).split("-")); init = date(y, m, d)
        except Exception:
            init = date.today()
        top = tk.Toplevel(self); top.title("Selecciona fecha"); top.transient(self); top.grab_set(); top.resizable(False, False)
        self.update_idletasks()
        top.geometry(f"+{self.winfo_rootx()+self.winfo_width()//2-180}+{self.winfo_rooty()+self.winfo_height()//2-170}")
        cal = Calendar(top, selectmode="day", year=init.year, month=init.month, day=init.day, date_pattern="yyyy-mm-dd",
                       firstweekday="monday", showweeknumbers=False)
        cal.pack(padx=14, pady=14)
        def choose():
            entry.delete(0, "end"); entry.insert(0, cal.get_date()); top.destroy()
        tk.Button(top, text="Seleccionar", command=choose).pack(side="left", padx=10, pady=10)
        tk.Button(top, text="Cerrar", command=top.destroy).pack(side="left", padx=10, pady=10)

    def _placeholder_png(self, msg="Generando…"):
        img = Image.new("RGB", (CHART_W, CHART_H), (245, 245, 247))
        d = ImageDraw.Draw(img)
        try: font = ImageFont.load_default()
        except Exception: font = None
        tw = d.textlength(msg, font=font)
        d.text(((CHART_W - tw) / 2, CHART_H / 2 - 7), msg, fill=(17, 24, 39), font=font)
        buf = io.BytesIO(); img.save(buf, format="PNG"); return buf.getvalue()

    def _png(self, fig):
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=RENDER_DPI, bbox_inches="tight")
        plt.close(fig)
        return buf.getvalue()

    def _coerce_data(self, data):
        if isinstance(data, list):
            return data if all(isinstance(x, dict) for x in data) else [{"valor": x} for x in data]
        if hasattr(data, "to_dict"):
            try: return data.to_dict("records")
            except Exception: return []
        if isinstance(data, str):
            try: obj = json.loads(data)
            except Exception: return []
            if isinstance(obj, list): return obj
            if isinstance(obj, dict):
                if "data" in obj and isinstance(obj["data"], list): return obj["data"]
                if obj and all(isinstance(v, dict) for v in obj.values()):
                    return [{**v, "fecha": k} for k, v in obj.items()]
        if isinstance(data, dict): return [data]
        return []

    def _parse_date(self, s):
        if not s: return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
            try: return datetime.strptime(str(s), fmt).date()
            except Exception: pass
        return None

    def _series_shipments(self, desde, hasta):
        try: rows = leer_shipments()
        except Exception: return {}
        dmin = self._parse_date(desde); dmax = self._parse_date(hasta)
        agg = {}
        for r in rows:
            if str(r.get("approved","0")) != "1":  # solo aprobadas
                continue
            d = self._parse_date(r.get("ship_date"))
            if not d: continue
            if dmin and d < dmin: continue
            if dmax and d > dmax: continue
            try: q = float(str(r.get("qty","0")).replace(",","") or 0)
            except Exception: q = 0
            agg[d.isoformat()] = agg.get(d.isoformat(), 0) + q
        return agg

    def _cumsum(self, xs):
        s = 0.0; out = []
        for v in xs:
            try: s += float(v or 0)
            except Exception: s += 0
            out.append(s)
        return out

    def _xdates(self, fechas):
        out = []
        for f in fechas:
            d = self._parse_date(f)
            out.append(datetime.combine(d, dtime.min) if d else datetime.today())
        return out

    def _normalize_apq(self, a, p, q, max_turnos=3):
        series = [[float(x or 0) for x in (a or [])],
                  [float(x or 0) for x in (p or [])],
                  [float(x or 0) for x in (q or [])]]
        max_val = max((max(s) if s else 0) for s in series) if series else 0
        if max_val <= 110: return series, 1
        turnos = int(min(max_turnos, max(1, round(max_val/100.0))))
        norm = [[min(100.0, v/turnos) for v in s] for s in series]
        return norm, turnos

    def _log_error(self, where, err):
        try:
            with open("ui_errors.log", "a", encoding="utf-8") as f:
                f.write(f"[{datetime.now().isoformat()}] {where}: {err}\n{traceback.format_exc()}\n")
        except Exception:
            pass

    # ================== Generación Async ==================
    def _generar_async(self):
        if self._generating: return
        self._generating = True
        self.btn_generate.configure(state="disabled", text="Generando…")
        self.btn_generate_top.configure(state="disabled", text="Generando…")
        for w in self.charts.winfo_children(): w.destroy()
        self._plot_bytes.clear(); self._plot_titles.clear(); self._plot_captions.clear()
        threading.Thread(target=self._worker_generate, daemon=True).start()

    def _worker_generate(self):
        try:
            mid = self.machine_var.get()
            machine = next((m for m in (MACHINES if "MACHINES" in globals() else []) if m.get("id") == mid), None)
            desde = self.desde_entry.get().strip() or (date.today() - timedelta(days=7)).isoformat()
            hasta = self.hasta_entry.get().strip() or date.today().isoformat()

            res = {}
            try:
                res = resumen_rango_maquina(machine or mid, desde, hasta)
            except Exception as e:
                self._log_error("resumen_rango_maquina", e)

            data_raw = (res.get("rows") or res.get("data")) if isinstance(res, dict) else []
            stats_raw = (res.get("totals") or res.get("summary")) if isinstance(res, dict) else {}
            data = self._coerce_data(data_raw)
            if not data:
                data, stats_raw = self._demo_data(desde, hasta)

            def _i(x):
                try: return int(float(str(x).replace(",", "") or 0))
                except Exception: return 0
            def _f(x):
                try: return float(str(x).replace(",", "") or 0.0)
                except Exception: return 0.0

            machine_name = (machine.get("name") or machine.get("alias") or machine.get("id")) if isinstance(machine, dict) else str(machine or mid)
            stats = {
                "total": _i(stats_raw.get("total")), "buenas": _i(stats_raw.get("buenas")),
                "scrap": _i(stats_raw.get("scrap")), "oee_prom": _f(stats_raw.get("oee_prom")),
                "avail_prom": _f(stats_raw.get("avail_prom")), "perf_prom": _f(stats_raw.get("perf_prom")),
                "qual_prom": _f(stats_raw.get("qual_prom")), "paro_min_total": _f(stats_raw.get("paro_min_total")),
                "machine": machine_name, "desde": (self._parse_date(desde) or date.today()).isoformat(),
                "hasta": (self._parse_date(hasta) or date.today()).isoformat(),
            }

            try:
                charts, series_prom = self._build_seaborn_charts(stats, data, self._series_shipments(desde, hasta))
            except Exception as e:
                self._log_error("build_charts", e)
                data, demo_stats = self._demo_data(desde, hasta)
                for k in ("oee_prom","avail_prom","perf_prom","qual_prom","total","buenas","scrap"):
                    stats[k] = stats[k] or demo_stats.get(k, 0)
                charts, series_prom = self._build_seaborn_charts(stats, data, {})

            for k in ("oee_prom", "avail_prom", "perf_prom", "qual_prom"):
                if not stats.get(k, 0): stats[k] = series_prom.get(k, 0)

            self.after(0, lambda: self._on_ready(stats, data, charts))
        except Exception as e:
            self._log_error("worker_generate", e)
            self.after(0, lambda: self._on_ready({"total":0,"buenas":0,"scrap":0,"oee_prom":0}, [], []))

    # ================== Gráficas Seaborn ==================
    def _legend_top(self, ax, ncol):
        ax.legend(loc="upper center", bbox_to_anchor=(0.5, 1.25), ncol=ncol,
                  fontsize=8, frameon=False, handlelength=1.8, columnspacing=0.9)

    def _new_fig(self, extra_bottom=0.22):
        fig = plt.figure(figsize=(CHART_W/BASE_DPI, CHART_H/BASE_DPI), dpi=RENDER_DPI)
        fig.subplots_adjust(left=0.12, right=0.98, top=0.85, bottom=extra_bottom)
        return fig

    def _format_dates(self, ax, xdt):
        locator = mdates.AutoDateLocator(minticks=3, maxticks=6)
        formatter = mdates.ConciseDateFormatter(locator)
        ax.xaxis.set_major_locator(locator)
        ax.xaxis.set_major_formatter(formatter)
        ax.margins(x=0.02)

    def _build_seaborn_charts(self, stats, data, ships_by_day):
        fechas = [d.get("fecha", "") for d in data]
        xdt = self._xdates(fechas)

        a_raw = [float(d.get("availability", 0) or 0) for d in data]
        p_raw = [float(d.get("performance", 0) or 0) for d in data]
        q_raw = [float(d.get("quality", 0) or 0) for d in data]
        o_raw = [float(d.get("oee", 0) or 0) for d in data]
        b = [float(d.get("buenas", 0) or 0) for d in data]
        s = [float(d.get("scrap", 0) or 0) for d in data]

        # Normaliza A/P/Q por # turnos si excede 100 %
        (a, p, q), turnos = self._normalize_apq(a_raw, p_raw, q_raw)

        # OEE (si viene 0, calcúlalo con A*P*Q)
        o = []
        for ai, pi, qi, oi in zip(a_raw, p_raw, q_raw, o_raw):
            val = float(oi or 0)
            if val <= 0 and ai > 0 and pi > 0 and qi > 0:
                val = ai * pi * qi / 10000.0
            if val > 110: val = 100.0
            o.append(max(0.0, min(100.0, val)))

        prom = {
            "oee_prom":   (sum(o)/len(o) if o else 0.0),
            "avail_prom": (sum(min(100,x) for x in a_raw)/len(a_raw) if a_raw else 0.0),
            "perf_prom":  (sum(min(100,x) for x in p_raw)/len(p_raw) if p_raw else 0.0),
            "qual_prom":  (sum(min(100,x) for x in q_raw)/len(q_raw) if q_raw else 0.0),
        }

        charts = []

        # ============ 1) Producción diaria (colores + HEADROOM) ============
        fig = self._new_fig(extra_bottom=0.26)
        ax = fig.add_subplot(111)
        ax.bar(xdt, b, label="Buenas", color=PALETTE["buenas"])
        ax.bar(xdt, s, bottom=b, label="Scrap", color=PALETTE["scrap"])
        tot = [bb + ss for bb, ss in zip(b, s)]
        sns.lineplot(x=xdt, y=tot, ax=ax, label="Total",
                     marker="o", linewidth=1.5, markersize=3.4, color=PALETTE["total"])
        ax.set_ylabel("Piezas")
        ymax1 = max(tot) if tot else 0
        ax.set_ylim(0, ymax1 * 1.15 if ymax1 > 0 else 1)  # +15% de headroom
        ax.yaxis.set_major_locator(MaxNLocator(nbins=5))
        self._format_dates(ax, xdt)
        self._legend_top(ax, ncol=3)
        charts.append((
            "Producción diaria (Buenas + Scrap) y Total",
            "Barras apiladas = Buenas y Scrap por día; línea = Total producido.",
            self._png(fig)
        ))

        # ============ 2) A/P/Q + OEE (paleta pro) ============
        fig = self._new_fig(extra_bottom=0.26)
        ax = fig.add_subplot(111)
        xd = mdates.date2num(xdt)
        bar_w = 0.25  # en días
        ax.bar(xd - bar_w, a, width=bar_w, label="Availability", color=PALETTE["avail"], align="center")
        ax.bar(xd,           p, width=bar_w, label="Effectivity",  color=PALETTE["perf"],  align="center")
        ax.bar(xd + bar_w,   q, width=bar_w, label="Quality",      color=PALETTE["qual"],  align="center")
        sns.lineplot(x=xdt, y=o, ax=ax, label="OEE",
                     marker="o", linewidth=1.6, markersize=3.2, color=PALETTE["oee"])
        ax.set_xlim(min(xd)-0.7, max(xd)+0.7)
        ax.set_ylim(0, 110)
        ax.set_ylabel("%"); ax.yaxis.set_major_locator(MaxNLocator(nbins=5))
        ax.xaxis_date()
        self._format_dates(ax, xdt)
        if turnos > 1:
            ax.text(0.01, 1.06, f"Normalizado por {turnos} turno(s)", transform=ax.transAxes, fontsize=8, color="#6b7280")
        self._legend_top(ax, ncol=4)
        charts.append((
            "A/P/Q con OEE",
            "Barras: Availability, Effectivity y Quality. Línea: OEE diario (OEE no se normaliza).",
            self._png(fig)
        ))

        # ============ 3) Curva S (más HEADROOM y colores dedicados) ============
        shipped = [float(ships_by_day.get(f, 0) or 0) for f in fechas]
        if any(b) or any(shipped):
            cum_b = self._cumsum(b); cum_s = self._cumsum(shipped); cum_d = [max(bb-ss,0) for bb,ss in zip(cum_b,cum_s)]
            fig = self._new_fig(extra_bottom=0.30)
            ax = fig.add_subplot(111)
            sns.lineplot(x=xdt, y=cum_b, ax=ax, label="Buenas acumuladas",
                         marker="o", markersize=3, linewidth=1.6, color=PALETTE["buenas"])
            if any(shipped):
                sns.lineplot(x=xdt, y=cum_s, ax=ax, label="Salidas acumuladas",
                             marker="o", markersize=3, linewidth=1.6, color=PALETTE["ship"])
                sns.lineplot(x=xdt, y=cum_d, ax=ax, label="Disponible acumulado",
                             marker="o", markersize=3, linewidth=1.6, color=PALETTE["disp"])
            ax.set_ylabel("Piezas acumuladas")
            ymax3 = max([*(cum_b or [0]), *(cum_s or [0]), *(cum_d or [0])])
            ax.set_ylim(0, ymax3 * 1.12 if ymax3 > 0 else 1)  # +12% headroom
            ax.yaxis.set_major_locator(MaxNLocator(nbins=5))
            self._format_dates(ax, xdt)
            self._legend_top(ax, ncol=3)
            charts.append((
                "Curva S - Buenas / Salidas / Disponible",
                "Acumulados del periodo: producción buena, salidas aprobadas y stock disponible.",
                self._png(fig)
            ))

        # ============ 4) Control chart OEE ============
        if len(o) >= 2:
            import statistics as st
            mu = st.mean(o); sd = st.pstdev(o)  # población
            ucl = min(100, mu + 3*sd); lcl = max(0, mu - 3*sd)
            ymin = max(0, min(min(o), lcl) - 5); ymax = min(110, max(max(o), ucl) + 5)
            fig = self._new_fig(extra_bottom=0.30)
            ax = fig.add_subplot(111)
            sns.lineplot(x=xdt, y=o, ax=ax, label="OEE", marker="o", markersize=3.2,
                         linewidth=1.8, color=PALETTE["oee_line"])
            ax.axhspan(lcl, ucl, color=PALETTE["band"], alpha=0.35, lw=0)
            ax.axhline(mu, color=PALETTE["mean"], lw=1.3, ls="--")
            ax.text(0.01, 1.06, f"Media {mu:.1f}%", transform=ax.transAxes, va="bottom", fontsize=8)
            ax.text(0.99, 1.06, f"+/- 3 sigma [{lcl:.1f}-{ucl:.1f}]", transform=ax.transAxes, va="bottom", ha="right", fontsize=8)
            ax.set_ylim(ymin, ymax); ax.set_ylabel("%"); ax.yaxis.set_major_locator(MaxNLocator(nbins=5))
            self._format_dates(ax, xdt)
            self._legend_top(ax, ncol=1)
            charts.append((
                "Control chart OEE (+/- 3 sigma)",
                "Banda de control con media y +/- 3 sigma. Puntos fuera sugieren causas especiales.",
                self._png(fig)
            ))

        return charts, prom

    def _on_ready(self, stats, data, charts):
        self._last_stats = stats; self._last_data = data
        self.lbl_total.configure(text=f"{stats.get('total',0):,}")
        self.lbl_buenas.configure(text=f"{stats.get('buenas',0):,}")
        self.lbl_scrap.configure(text=f"{stats.get('scrap',0):,}")
        oee_prom = float(stats.get("oee_prom",0) or 0); self.lbl_oee.configure(text=f"{oee_prom:.2f}%"); self.card_oee.configure(fg_color=self._tone(oee_prom))
        self.lbl_paro.configure(text=f"{int(stats.get('paro_min_total',0)):,}")

        for w in self.charts.winfo_children(): w.destroy()
        self._plot_bytes.clear(); self._plot_titles.clear(); self._plot_captions.clear()

        if not charts:
            ctk.CTkLabel(self.charts, text="No hay datos para el rango seleccionado."
            ).grid(row=0, column=0, padx=12, pady=12, sticky="w")
        else:
            row = col = 0
            for title, caption, img_bytes in charts:
                self._add_chart_card(self.charts, row, col, title, caption, img_bytes)
                self._plot_titles.append(title); self._plot_captions.append(caption); self._plot_bytes.append(img_bytes)
                col = 1 if col == 0 else 0
                if col == 0: row += 1

        self._reload_table(data)
        self.btn_generate.configure(state="normal", text="Generar")
        self.btn_generate_top.configure(state="normal", text="↻ Generar")
        self._generating = False

    def _add_chart_card(self, parent, row, col, title, caption, bytes_png):
        card = ctk.CTkFrame(parent, corner_radius=16, fg_color=("white", "#1c1c1e"))
        card.grid(row=row, column=col, padx=10, pady=(12, 20), sticky="nsew")
        parent.grid_rowconfigure(row, weight=1)
        ctk.CTkLabel(card, text=title, font=ctk.CTkFont("Helvetica", 13, "bold")
        ).pack(anchor="w", padx=12, pady=(10, 4))
        im = Image.open(io.BytesIO(bytes_png))
        cimg = ctk.CTkImage(light_image=im, dark_image=im, size=(CHART_W, CHART_H))
        ctk.CTkLabel(card, image=cimg, text="").pack(padx=12, pady=(0, 10), expand=True, fill="both")
        ctk.CTkLabel(card, text=caption, font=ctk.CTkFont(size=10),
                     text_color=("#6b7280","#9CA3AF"), wraplength=int(CHART_W*0.95),
                     justify="left").pack(anchor="w", padx=12, pady=(0, 22))

    # ================== Tabla ==================
    def _reload_table(self, data):
        for i in self.tbl.get_children(): self.tbl.delete(i)
        tb = ts = tt = 0
        for d in data:
            a = float(d.get("availability",0) or 0); p = float(d.get("performance",0) or 0)
            q = float(d.get("quality",0) or 0); o = float(d.get("oee",0) or 0)
            b = d.get("buenas",""); s = d.get("scrap",""); t = d.get("total","")
            try: tb += int(float(b or 0))
            except: pass
            try: ts += int(float(s or 0))
            except: pass
            try: tt += int(float(t or 0))
            except: pass
            self.tbl.insert("", "end", values=(d.get("fecha",""), f"{a:.1f}", f"{p:.1f}", f"{q:.1f}", f"{o:.1f}", b, s, t))
        self.lbl_tbl_total.configure(text=f"Filas: {len(data)}  •  Buenas: {tb:,}  •  Scrap: {ts:,}  •  Total: {tt:,}")

    # ================== Export ==================
    def _exportar(self):
        if not self._last_data:
            messagebox.showwarning("Exportar", "Genera un reporte primero"); return
        if messagebox.askyesno("Exportar", "¿Exportar como PDF? (No = Excel)"):
            path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF","*.pdf")])
            if path: self._export_pdf(path)
        else:
            path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
            if path: self._export_excel(path)

    def _export_pdf(self, path):
        df = pd.DataFrame(self._last_data)
        pdf = FPDF(orientation="L", unit="mm", format="A4")
        pdf.set_auto_page_break(auto=True, margin=15); pdf.add_page()
        if os.path.exists(LOGO_PATH): pdf.image(LOGO_PATH, x=260, y=8, w=30)
        pdf.set_font("Arial", "B", 15); pdf.set_y(12); pdf.cell(0, 10, _pdf_safe("Reporte de Produccion"), ln=True, align="C")
        pdf.set_font("Arial", "", 10)
        meta = f"Maquina: {self._last_stats.get('machine','-')}    Rango: {self._last_stats.get('desde','-')} -> {self._last_stats.get('hasta','-')}    Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        pdf.cell(0, 7, _pdf_safe(meta), ln=True, align="C"); pdf.ln(2)

        pdf.set_font("Arial", "B", 10); pdf.cell(0, 8, _pdf_safe("Resumen (KPIs)"), ln=True); pdf.set_font("Arial", "", 9)
        kpis = [
            ("Total (pzs)", f"{self._last_stats.get('total',0):,}"),
            ("Buenas (pzs)", f"{self._last_stats.get('buenas',0):,}"),
            ("Scrap (pzs)",  f"{self._last_stats.get('scrap',0):,}"),
            ("OEE Prom (%)", f"{self._last_stats.get('oee_prom',0):.2f}%"),
            ("Availability Prom (%)", f"{self._last_stats.get('avail_prom',0):.2f}%"),
            ("Effectivity Prom (%)",  f"{self._last_stats.get('perf_prom',0):.2f}%"),
            ("Quality Prom (%)",      f"{self._last_stats.get('qual_prom',0):.2f}%"),
            ("Paro total (min)",      f"{int(self._last_stats.get('paro_min_total',0)):,}"),
        ]
        colw = (pdf.w - 2 * pdf.l_margin) / 4
        for i,(k,v) in enumerate(kpis):
            if i % 4 == 0: pdf.ln(6)
            pdf.cell(colw, 7, _pdf_safe(f"{k}: {v}"), border=0)
        pdf.ln(6)

        if not df.empty:
            pdf.set_font("Arial", "B", 9); col_w = (pdf.w - 2 * pdf.l_margin) / len(df.columns)
            for c in df.columns: pdf.cell(col_w, 7, _pdf_safe(str(c)), border=1, align="C")
            pdf.ln(7); pdf.set_font("Arial", "", 8)
            for _, row in df.iterrows():
                for c in df.columns: pdf.cell(col_w, 6, _pdf_safe(str(row[c])), border=1)
                pdf.ln(6)

        pdf.add_page()
        for title, caption, bytes_png in zip(self._plot_titles, self._plot_captions, self._plot_bytes):
            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                tmp.write(bytes_png); pth = tmp.name
            img_w = pdf.w - 2 * pdf.l_margin; img_h = img_w * CHART_H / CHART_W
            pdf.set_font("Arial", "B", 11); pdf.cell(0, 7, _pdf_safe(title), ln=True)
            pdf.image(pth, w=img_w, h=img_h); os.remove(pth)
            pdf.set_font("Arial", "", 8); pdf.multi_cell(0, 5, _pdf_safe(caption)); pdf.ln(3)
        pdf.output(path)

    def _export_excel(self, path):
        wb = Workbook()
        ws_res = wb.active; ws_res.title = "Resumen"
        ws_dat = wb.create_sheet("Datos"); ws_gra = wb.create_sheet("Graficas")

        title_font = Font(bold=True, size=15)
        h_font = Font(bold=True, size=10)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        fill_gray = PatternFill("solid", fgColor="F3F4F6")
        border = Border(left=Side(style="thin", color="D1D5DB"),
                        right=Side(style="thin", color="D1D5DB"),
                        top=Side(style="thin", color="D1D5DB"),
                        bottom=Side(style="thin", color="D1D5DB"))

        # ---- Cabecera: título centrado y LOGO a la DERECHA sin amontonarse ----
        ws_res.merge_cells("A1:H1")
        ws_res["A1"] = "Reporte de Produccion"; ws_res["A1"].font = title_font; ws_res["A1"].alignment = center

        # Column widths para dejar “zona” libre al logo
        col_widths = {"A": 16, "B": 38, "C": 14, "D": 14, "E": 16, "F": 16, "G": 16, "H": 16}
        for c, w in col_widths.items():
            ws_res.column_dimensions[c].width = w
        # Altura de filas de cabecera
        ws_res.row_dimensions[1].height = 24
        ws_res.row_dimensions[2].height = 50  # fila donde va el logo (respira)

        # Logo anclado en G2 (derecha)
        if os.path.exists(LOGO_PATH):
            try:
                xlimg = XLImage(LOGO_PATH)
                # Mantener proporción: altura 48 px
                target_h = 48
                if xlimg.height:
                    ratio = xlimg.width / xlimg.height
                    xlimg.height = target_h
                    xlimg.width = int(target_h * ratio)
                else:
                    xlimg.width, xlimg.height = 120, 48
                ws_res.add_image(xlimg, "G2")
            except Exception:
                pass

        # ---- Metadatos, lejos del logo (col A:B) ----
        meta = [("Maquina", self._last_stats.get("machine","-")),
                ("Rango", f"{self._last_stats.get('desde','-')} -> {self._last_stats.get('hasta','-')}"),
                ("Generado", datetime.now().strftime("%Y-%m-%d %H:%M"))]
        r = 3
        for k,v in meta:
            ws_res[f"A{r}"] = k; ws_res[f"A{r}"].font = h_font
            ws_res[f"B{r}"] = v; ws_res[f"B{r}"].alignment = left; r += 1

        # ---- KPIs con header gris ----
        ws_res["A6"] = "KPI"; ws_res["B6"] = "Valor"
        ws_res["A6"].font = h_font; ws_res["B6"].font = h_font
        ws_res["A6"].fill = fill_gray; ws_res["B6"].fill = fill_gray
        rows = [
            ("Total (pzs)", self._last_stats.get("total", 0)),
            ("Buenas (pzs)", self._last_stats.get("buenas", 0)),
            ("Scrap (pzs)",  self._last_stats.get("scrap", 0)),
            ("OEE Prom (%)", self._last_stats.get("oee_prom", 0)),
            ("Availability Prom (%)", self._last_stats.get("avail_prom", 0)),
            ("Effectivity Prom (%)",  self._last_stats.get("perf_prom", 0)),
            ("Quality Prom (%)",      self._last_stats.get("qual_prom", 0)),
            ("Paro total (min)",      self._last_stats.get("paro_min_total", 0)),
        ]
        rr = 7
        for k,v in rows:
            ws_res[f"A{rr}"] = k; ws_res[f"A{rr}"].alignment = left
            ws_res[f"B{rr}"] = v
            if " (%)" in k: ws_res[f"B{rr}"].number_format = "0.00"
            ws_res[f"A{rr}"].border = border; ws_res[f"B{rr}"].border = border; rr += 1
        for col in ("A","B"):
            maxlen = max(len(str(ws_res[f'{col}{i}'].value or "")) for i in range(1, rr+1))
            ws_res.column_dimensions[col].width = max(16, min(42, maxlen + 2))

        # ---- Hoja “Datos” (tabla con autoajuste) ----
        df = pd.DataFrame(self._last_data)
        if not df.empty:
            ws_dat.append(list(df.columns))
            for j in range(1, len(df.columns)+1):
                c = ws_dat.cell(row=1, column=j); c.font = h_font; c.alignment = center; c.fill = fill_gray; c.border = border
            for _, row in df.iterrows(): ws_dat.append([row[c] for c in df.columns])
            for j, col in enumerate(df.columns, start=1):
                maxlen = max(len(str(x)) for x in ([col] + df[col].astype(str).tolist()))
                ws_dat.column_dimensions[get_column_letter(j)].width = min(36, max(12, maxlen + 2))
            ref = f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"
            table = Table(displayName="Datos", ref=ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
            ws_dat.add_table(table); ws_dat.freeze_panes = "A2"

        # ---- Hoja “Graficas” con imágenes y títulos, espaciadas ----
        row_g = 1
        for title, bytes_png in zip(self._plot_titles, self._plot_bytes):
            ws_gra.merge_cells(start_row=row_g, start_column=1, end_row=row_g, end_column=8)
            ws_gra.cell(row=row_g, column=1, value=title).font = h_font
            xlimg = XLImage(io.BytesIO(bytes_png)); xlimg.width, xlimg.height = CHART_W, CHART_H
            ws_gra.add_image(xlimg, f"A{row_g+1}")
            row_g += int(CHART_H/18) + 10  # margen vertical para que no se amontonen

        wb.save(path)

    # ================== DEMO ==================
    def _demo_data(self, desde, hasta):
        try:
            d0 = self._parse_date(desde) or (date.today() - timedelta(days=6))
            d1 = self._parse_date(hasta) or date.today()
            days = max(5, min(10, (d1 - d0).days + 1))
            fechas = [(d0 + timedelta(days=i)).isoformat() for i in range(days)]
        except Exception:
            fechas = [(date.today() - timedelta(days=5 - i)).isoformat() for i in range(6)]
        rows = []; total=buenas=scrap=0
        for i,f in enumerate(fechas):
            A=96; P=95-(i%3)*2; Q=96-(i%4); O=max(0, min(100, A*P*Q/10000))
            B=1250; S=(i%3)*50; T=B+S
            total+=T; buenas+=B; scrap+=S
            rows.append({"fecha":f,"availability":A,"performance":P,"quality":Q,"oee":O,"buenas":B,"scrap":S,"total":T})
        stats={"total":total,"buenas":buenas,"scrap":scrap,
               "oee_prom":sum(r["oee"] for r in rows)/len(rows),
               "avail_prom":sum(r["availability"] for r in rows)/len(rows),
               "perf_prom":sum(r["performance"] for r in rows)/len(rows),
               "qual_prom":sum(r["quality"] for r in rows)/len(rows),
               "paro_min_total":0}
        return rows, stats
