# -*- coding: utf-8 -*-
# Recetas de Máquina (layout estilo SELOGICA) — versiones + historial con grilla y export
from .base import *  # ctk, tk, ttk, messagebox, filedialog, BASE_DIR
import os, json, csv, sys, re
from datetime import datetime

# ---------- Paths ----------
RECIPES_DIR = os.path.join(BASE_DIR, "machine_recipes")
HISTORY_CSV = os.path.join(RECIPES_DIR, "_history.csv")
os.makedirs(RECIPES_DIR, exist_ok=True)

DATA_SCHEMA_VERSION = 2

# ---------- Excel helpers ----------
EXCEL_MAP = {
    "program": "B3", "mould_desig": "B4", "material": "B5",
    "date_of_entry": "E3", "cavities": "E4", "machine": "E5",
    "cycle_time_s": "E7", "injection_time_s": "E8", "holding_press_time_s": "E9",
    "rem_cooling_time_s": "E10", "dosage_time_s": "E11", "screw_stroke_mm": "E12",
    "mould_stroke_mm": "E13", "ejector_stroke_mm": "E14", "shot_weight_g": "E15",
    "plasticising_flow_kgh": "E16", "dosage_capacity_gs": "E17", "dosage_volume_ccm": "E18",
    "material_cushion_ccm": "E19", "max_inj_pressure_bar": "E20",
}
# Nota: para cubrir Injection / Plasticizing / Holding / Temperatures /
# Movements / Clamping, amplía este mapa con las celdas correspondientes.

NUM_FIELDS = {
    "cycle_time_s","injection_time_s","holding_press_time_s","rem_cooling_time_s",
    "dosage_time_s","screw_stroke_mm","mould_stroke_mm","ejector_stroke_mm",
    "shot_weight_g","plasticising_flow_kgh","dosage_capacity_gs","dosage_volume_ccm",
    "material_cushion_ccm","max_inj_pressure_bar","screw_d_mm","pcs_1",
    "inj_press_lim_1","inj_press_lim_2","inj_press_lim_3",
    "inj_speed_1","inj_speed_2","inj_speed_3",
    "inj_end_stage_mm_1","inj_end_stage_mm_2","inj_end_stage_mm_3",
    "inj_flow_1","inj_flow_2","inj_flow_3",
    "inj_end_stage_ccm_1","inj_end_stage_ccm_2","inj_end_stage_ccm_3",
    "plast_screw_speed","plast_back_pressure","plast_end_stage_ccm",
    "hp_time_1","hp_time_2","hp_time_3",
    "hp_press_1","hp_press_2","hp_press_3","hp_press_4",
    "temp_c1","temp_c2","temp_c3","temp_c4","temp_c5",
    "tol_c1","tol_c2","tol_c3","tol_c4","tol_c5",
    "feed_yoke_temp","lower_enable_tol","upper_switch_off_tol",
    "open_end_mm_1","open_end_mm_2","open_end_mm_3",
    "open_speed_1","open_speed_2","open_speed_3",
    "open_force_1","open_force_2","open_force_3",
    "close_end_mm_1","close_end_mm_2","close_end_mm_3","close_end_mm_4",
    "close_speed_1","close_speed_2","close_speed_3","close_speed_4",
    "close_force_1","close_force_2","close_force_3",
    "mould_closed_kn",
}

# ---------- Template Excel (formato predefinido) ----------
# Intenta estas rutas en orden; usa la primera que exista.
TEMPLATE_CANDIDATES = [
    os.path.join(BASE_DIR, "formatorecta.xlsx"),
    os.path.join(RECIPES_DIR, "formatorecta.xlsx"),
    os.path.join(BASE_DIR, "assets", "formatorecta.xlsx"),
    "/mnt/data/formatorecta.xlsx",  # por si ya lo tienes ahí
]

def _find_excel_template():
    for p in TEMPLATE_CANDIDATES:
        if os.path.exists(p):
            return p
    return None

def _to_excel_value(key: str, raw_val):
    """
    Convierte a número si el campo es numérico. Si no se puede, deja texto.
    Limpia caracteres ilegales para Excel.
    """
    s = "" if raw_val is None else str(raw_val).strip()
    if key in NUM_FIELDS:
        # intenta número (permite coma como separador de miles y punto decimal)
        try:
            s2 = s.replace(",", "")
            if s2 == "":
                return None
            v = float(s2)
            # si es entero puro, déjalo como int
            return int(v) if abs(v - int(v)) < 1e-12 else v
        except Exception:
            # si no parsea como número, escribe como texto para no perder el dato
            pass
    # texto (sanitizado)
    return _safe_excel(s)

def _export_snapshot_to_template(snapshot: dict, out_path: str, template_path: str, sheet_name: str | None = None):
    """
    Abre la plantilla y escribe los valores según EXCEL_MAP en la hoja indicada (o activa).
    Mantiene estilos/formatos del template.
    """
    from openpyxl import load_workbook

    wb = load_workbook(template_path, data_only=False, keep_vba=False)
    ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active

    # Escribe todos los campos mapeados
    for fid, cell_addr in EXCEL_MAP.items():
        val = snapshot.get(fid, "")
        ws[cell_addr].value = _to_excel_value(fid, val)

    # Si quieres estampar fecha/hora o versión en alguna celda, puedes hacerlo aquí (opcional):
    # ej: ws["H2"].value = datetime.now().strftime("%Y-%m-%d %H:%M")

    wb.save(out_path)

# ---------- FS helpers ----------
def _safe_id(s: str) -> str:
    return str(s).replace("/", "_").replace("\\", "_").strip()

def _path_json(mold_id: str) -> str:
    return os.path.join(RECIPES_DIR, f"{_safe_id(mold_id)}.json")

def _versions_dir(mold_id: str) -> str:
    d = os.path.join(RECIPES_DIR, _safe_id(mold_id), "_versions")
    os.makedirs(d, exist_ok=True)
    return d

def _list_versions(mold_id: str):
    d = _versions_dir(mold_id)
    items = []
    rx = re.compile(r"^v(\d{3})\.json$")
    for fname in os.listdir(d):
        m = rx.match(fname)
        if m:
            items.append((m.group(0).replace(".json",""), os.path.join(d, fname)))
    items.sort(key=lambda t: t[0])
    return items

def _next_version_name(mold_id: str) -> str:
    items = _list_versions(mold_id)
    if not items: return "v001"
    last = items[-1][0]
    return f"v{int(last[1:])+1:03d}"

def _load_json(mold_id: str) -> dict:
    p = _path_json(mold_id)
    if os.path.exists(p):
        try:
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def _save_json(mold_id: str, data: dict):
    payload = dict(data)
    payload["_meta"] = {
        "schema": DATA_SCHEMA_VERSION,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    with open(_path_json(mold_id), "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

def _save_version_snapshot(mold_id: str, data: dict, usuario: str, motivo: str, diffs_text: str) -> str:
    ver = _next_version_name(mold_id)
    snap = dict(data)
    snap["_meta"] = {
        "schema": DATA_SCHEMA_VERSION,
        "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "usuario": usuario,
        "motivo": motivo,
        "version": ver,
        "diffs": diffs_text,
    }
    out = os.path.join(_versions_dir(mold_id), f"{ver}.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(snap, f, ensure_ascii=False, indent=2)
    return ver

def _load_version_snapshot(mold_id: str, version: str) -> dict:
    path = os.path.join(_versions_dir(mold_id), f"{version}.json")
    if not os.path.exists(path): return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def _cast_numeric(val: str) -> str:
    if val is None: return ""
    s = str(val).strip().replace(",", ".")
    if s == "": return ""
    try:
        num = float(s)
        return str(int(num)) if abs(num - int(num)) < 1e-9 else f"{num:.3f}".rstrip("0").rstrip(".")
    except Exception:
        return str(val).strip()

def _ascii(s: str) -> str:
    """Evita errores latin-1 al exportar PDFs en algunos backends."""
    if s is None: return ""
    return (str(s).replace("\u2014","-").encode("latin-1","replace").decode("latin-1"))

# ---------- Sanitizers para export ----------
try:
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
except Exception:
    ILLEGAL_CHARACTERS_RE = None

def _safe_excel(val: str) -> str:
    s = "" if val is None else str(val)
    if ILLEGAL_CHARACTERS_RE:
        s = ILLEGAL_CHARACTERS_RE.sub("", s)
    return s

def _safe_pdf(val: str) -> str:
    # Latin-1 friendly + sin saltos raros
    return _ascii(val).replace("\r", " ").replace("\n", " ")

# ============================================================
#                          VIEW
# ============================================================
class MachineRecipesView(ctk.CTkFrame):
    """Panel de recetas de máquina con layout SELOGICA + historial/versiones."""

    def __init__(self, master, app):
        super().__init__(master, fg_color="transparent")
        self.app = app
        self.vars = {}          # id -> StringVar
        self.current_mold = ""
        self._dirty = False
        self._preview_version = None
        self._status_var = tk.StringVar(value="")
        self._motivo_ph = (
            "Formato sugerido: [Qué cambiaste] -> [Por qué] -> [Resultado esperado].\n"
            "Ej.: Aumenté HP de 60->70 bar para mejorar compacción; espero eliminar porosidad en esquina A."
        )
        self._build()

    # --------------------------- UI ---------------------------
    def _build(self):
        # Header
        header = ctk.CTkFrame(self, corner_radius=0, fg_color=("white", "#0e1117"))
        header.pack(fill="x", side="top")
        left = ctk.CTkFrame(header, fg_color="transparent"); left.pack(side="left", padx=16, pady=10)
        ctk.CTkButton(left, text="← Menú", width=110, corner_radius=10,
                      fg_color="#E5E7EB", text_color="#111", hover_color="#D1D5DB",
                      command=self._back_menu_guard).pack(side="left", padx=(0, 10))
        title = ctk.CTkFrame(left, fg_color="transparent"); title.pack(side="left")
        ctk.CTkLabel(title, text="Recetas de Máquina", font=ctk.CTkFont("Helvetica", 22, "bold")).pack(anchor="w")
        ctk.CTkLabel(title, text="Layout estilo SELOGICA • Historial y versiones",
                     text_color=("#6b7280", "#9CA3AF")).pack(anchor="w")

        # Toolbar
        tools = ctk.CTkFrame(self, fg_color="transparent"); tools.pack(fill="x", padx=16, pady=(8, 0))
        opciones = ["— Selecciona —"] + list(getattr(self.app, "recipe_map", {}).keys())
        self.mold_var = tk.StringVar(value=opciones[0])
        self.mold_menu = ctk.CTkOptionMenu(tools, values=opciones, variable=self.mold_var, width=260,
                                           command=lambda *_: self._on_pick_mold())
        self.mold_menu.pack(side="left", padx=(0, 8))
        ctk.CTkButton(tools, text="📥 Importar Excel", command=self._import_excel).pack(side="left", padx=6)
        ctk.CTkButton(tools, text="🕓 Historial / Versiones", command=self._open_history).pack(side="left", padx=6)
        ctk.CTkButton(tools, text="📂 Carpeta de recetas",
                      command=lambda: self._open_folder(RECIPES_DIR)).pack(side="left", padx=6)

        # === DOCUMENTAR CAMBIO ===
        doc_card = ctk.CTkFrame(self, corner_radius=12, fg_color=("white", "#111827"))
        doc_card.pack(fill="x", padx=16, pady=(10, 0))
        topbar = ctk.CTkFrame(doc_card, fg_color="transparent"); topbar.pack(fill="x", padx=12, pady=(10, 4))
        ctk.CTkLabel(topbar, text="Documentar cambio", font=ctk.CTkFont(size=14, weight="bold")).pack(side="left")
        actions = ctk.CTkFrame(topbar, fg_color="transparent"); actions.pack(side="right")
        ctk.CTkButton(actions, text="💾 Guardar (Ctrl+S)", command=self._save).pack(side="left")
        ctk.CTkButton(actions, text="🗑", fg_color="#ef4444", hover_color="#dc2626", width=44,
                      command=self._clear_all).pack(side="left", padx=(8,0))
        body = ctk.CTkFrame(doc_card, fg_color="transparent"); body.pack(fill="both", padx=12, pady=(0, 12))
        self.motivo_txt = ctk.CTkTextbox(body, height=110)
        self.motivo_txt.pack(fill="x", expand=False)
        self._init_motivo_placeholder()
        btns = ctk.CTkFrame(body, fg_color="transparent"); btns.pack(fill="x", pady=(6,0))
        ctk.CTkButton(btns, text="Ejemplo", width=80, command=self._insert_motivo_example).pack(side="left")

        # status
        ctk.CTkLabel(self, textvariable=self._status_var,
                     text_color=("#6b7280", "#9CA3AF"), anchor="w").pack(fill="x", padx=16, pady=(4,0))

        # CONTENIDO con scroll (formulario)
        self.scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.scroll.pack(fill="both", expand=True, padx=16, pady=16)

        sheet = ctk.CTkFrame(self.scroll, corner_radius=14, fg_color=("white", "#111827"))
        sheet.pack(fill="x", padx=2, pady=2)
        sheet.grid_columnconfigure(0, weight=3)
        sheet.grid_columnconfigure(1, weight=2)

        hdr = self._card(sheet, "Parameter overview (Cabecera)", row=0, col=0)
        self._header_grid(hdr)
        key = self._card(sheet, "Key data", row=0, col=1)
        self._keydata_table(key)
        blocks = self._card(sheet, "", row=1, col=0, col_span=2)
        self._full_left_side(blocks)

        self._register_shortcuts_once()

    # ----- small UI builders -----
    def _card(self, parent, title, row=0, col=0, col_span=1):
        card = ctk.CTkFrame(parent, corner_radius=14, fg_color=("white", "#111827"))
        card.grid(row=row, column=col, columnspan=col_span, padx=8, pady=8, sticky="nsew")
        if title:
            ctk.CTkLabel(card, text=title, font=ctk.CTkFont("Helvetica", 13, "bold")).pack(anchor="w", padx=10, pady=(10, 0))
            ctk.CTkFrame(card, height=1, fg_color=("#E5E7EB", "#2B2B2B")).pack(fill="x", padx=10, pady=(6, 10))
        return card

    def _lbl(self, parent, text, **grid):
        l = ctk.CTkLabel(parent, text=text); l.grid(**grid); return l

    def _ent(self, parent, fid, w=110, justify="center", **grid):
        v = self.vars.setdefault(fid, tk.StringVar())
        e = ctk.CTkEntry(parent, textvariable=v, width=w, justify=justify)
        e.grid(**grid)
        v.trace_add("write", lambda *_: self._mark_dirty())
        return e

    def _header_grid(self, parent):
        g = ctk.CTkFrame(parent, fg_color="transparent"); g.pack(fill="x", padx=10, pady=10)
        for i in range(6): g.grid_columnconfigure(i, weight=1)
        self._lbl(g, "Program",        row=0, column=0, sticky="w", padx=4, pady=4)
        self._ent(g, "program",        row=0, column=1, sticky="ew", padx=4, pady=4)
        self._lbl(g, "Date of entry:", row=0, column=2, sticky="w", padx=4, pady=4)
        self._ent(g, "date_of_entry",  row=0, column=3, sticky="ew", padx=4, pady=4)
        self._lbl(g, "Cavities",       row=0, column=4, sticky="w", padx=4, pady=4)
        self._ent(g, "cavities",       row=0, column=5, sticky="ew", padx=4, pady=4)
        self._lbl(g, "Mould desig.",   row=1, column=0, sticky="w", padx=4, pady=4)
        self._ent(g, "mould_desig",    row=1, column=1, sticky="ew", padx=4, pady=4)
        self._lbl(g, "Machine",        row=1, column=2, sticky="w", padx=4, pady=4)
        self._ent(g, "machine",        row=1, column=3, sticky="ew", padx=4, pady=4)
        self._lbl(g, "Material",       row=2, column=0, sticky="w", padx=4, pady=4)
        self._ent(g, "material",       row=2, column=1, sticky="ew", padx=4, pady=4)

    def _keydata_table(self, parent):
        grid = ctk.CTkFrame(parent, fg_color="transparent"); grid.pack(fill="x", padx=10, pady=10)
        for i in range(2): grid.grid_columnconfigure(i, weight=1)
        rows = [
            ("cycle_time_s",           "Cycle time [s]"),
            ("injection_time_s",       "Injection time [s]"),
            ("holding_press_time_s",   "Holding press. time [s]"),
            ("rem_cooling_time_s",     "Rem. cooling time [s]"),
            ("dosage_time_s",          "Dosage time [s]"),
            ("screw_stroke_mm",        "Screw stroke [mm]"),
            ("mould_stroke_mm",        "Mould stroke [mm]"),
            ("ejector_stroke_mm",      "Ejector stroke [mm]"),
            ("shot_weight_g",          "Shot weight [g]"),
            ("plasticising_flow_kgh",  "Plasticising flow [kg/h]"),
            ("dosage_capacity_gs",     "Dosage capacity [g/s]"),
            ("dosage_volume_ccm",      "Dosage volume [ccm]"),
            ("material_cushion_ccm",   "Material cushion [ccm]"),
            ("max_inj_pressure_bar",   "max. inj. pressure [bar]"),
        ]
        for r, (fid, label) in enumerate(rows):
            self._lbl(grid, label, row=r, column=0, sticky="w", padx=4, pady=4)
            self._ent(grid, fid, row=r, column=1, sticky="ew", padx=4, pady=4)

    def _full_left_side(self, parent):
        ctk.CTkLabel(parent, text="Injection unit", font=ctk.CTkFont("Helvetica", 12, "bold")
                     ).pack(anchor="center", pady=(4, 0))
        inj_unit = ctk.CTkFrame(parent, fg_color="transparent"); inj_unit.pack(fill="x", padx=10, pady=(0, 8))
        inj_unit.grid_columnconfigure((0,1,2,3,4), weight=1)
        self._lbl(inj_unit, "Screw Ø [mm]", row=0, column=0, sticky="e", padx=6, pady=4)
        self._ent(inj_unit, "screw_d_mm", w=80, row=0, column=1, sticky="w", padx=4, pady=4)
        self._lbl(inj_unit, "Pcs. 1", row=0, column=3, sticky="e", padx=6, pady=4)
        self._ent(inj_unit, "pcs_1", w=80, row=0, column=4, sticky="w", padx=4, pady=4)

        self._section_table(parent, "Injection", [
            ("Injection press. limiting [bar]", ["inj_press_lim_1", "inj_press_lim_2", "inj_press_lim_3"]),
            ("Injection speed [mm/s]",          ["inj_speed_1", "inj_speed_2", "inj_speed_3"]),
            ("End of stage [mm]",               ["inj_end_stage_mm_1", "inj_end_stage_mm_2", "inj_end_stage_mm_3"]),
            ("Injection flow [ccm/s]",          ["inj_flow_1", "inj_flow_2", "inj_flow_3"]),
            ("End of stage [ccm]",              ["inj_end_stage_ccm_1", "inj_end_stage_ccm_2", "inj_end_stage_ccm_3"]),
        ])

        self._section_table(parent, "Plasticizing (St.1)", [
            ("Screw speed [m/min]",             ["plast_screw_speed"]),
            ("Back pressure [bar]",             ["plast_back_pressure"]),
            ("End of stage [ccm]",              ["plast_end_stage_ccm"]),
        ])

        self._section_table(parent, "Holding pressure (Pcs.2)", [
            ("Time [s]",                        ["hp_time_1", "hp_time_2", "hp_time_3"]),
            ("Pressure [bar]",                  ["hp_press_1", "hp_press_2", "hp_press_3", "hp_press_4"]),
        ])

        self._section_table(parent, "Temperatures (1..5)", [
            ("Cylinder temp. [°C]",             ["temp_c1", "temp_c2", "temp_c3", "temp_c4", "temp_c5"]),
            ("Tolerances [°C]",                 ["tol_c1", "tol_c2", "tol_c3", "tol_c4", "tol_c5"]),
            ("Feed yoke temperature [°C]",      ["feed_yoke_temp"]),
            ("Lower enable tol. [°C]",          ["lower_enable_tol"]),
            ("Upper switch-off tol. [°C]",      ["upper_switch_off_tol"]),
        ])

        self._section_table(parent, "Mould movements — Opening (St.1 / St.2 / St.3)", [
            ("End of stage [mm]",               ["open_end_mm_1", "open_end_mm_2", "open_end_mm_3"]),
            ("Speed [mm/s]",                    ["open_speed_1", "open_speed_2", "open_speed_3"]),
            ("Force [kN]",                      ["open_force_1", "open_force_2", "open_force_3"]),
        ])

        self._section_table(parent, "Mould movements — Closing (St.1 / St.2 / St.3 / An. HD)", [
            ("End of stage [mm]",               ["close_end_mm_1", "close_end_mm_2", "close_end_mm_3", "close_end_mm_4"]),
            ("Speed [mm/s]",                    ["close_speed_1", "close_speed_2", "close_speed_3", "close_speed_4"]),
            ("Force [kN]",                      ["close_force_1", "close_force_2", "close_force_3"]),
        ])

        self._section_table(parent, "Clamping", [
            ("Mould closed [kN]",               ["mould_closed_kn"]),
        ])

    def _section_table(self, parent, title, rows):
        card = ctk.CTkFrame(parent, corner_radius=12, fg_color=("white", "#111a27"))
        card.pack(fill="x", padx=10, pady=8)
        ctk.CTkLabel(card, text=title, font=ctk.CTkFont("Helvetica", 12, "bold")).pack(anchor="w", padx=10, pady=(8, 6))
        grid = ctk.CTkFrame(card, fg_color="transparent"); grid.pack(fill="x", padx=10, pady=(0,10))
        max_cols = max(len(cols) for _, cols in rows)
        for i in range(max_cols + 1): grid.grid_columnconfigure(i, weight=1)
        for r, (label, ids) in enumerate(rows):
            self._lbl(grid, label, row=r, column=0, sticky="w", padx=4, pady=4)
            for c, fid in enumerate(ids, start=1):
                self._ent(grid, fid, w=90, row=r, column=c, sticky="ew", padx=4, pady=4)

    # --------------------------- Eventos básicos ---------------------------
    def _register_shortcuts_once(self):
        root = self.winfo_toplevel()
        if getattr(root, "_recipes_shortcuts_ok", False): return
        root.bind("<Control-s>", lambda e: (self._save(), "break"), add="+")
        root._recipes_shortcuts_ok = True

    def _init_motivo_placeholder(self):
        # placeholder gris
        self._motivo_placeholder_on = True
        self.motivo_txt.delete("1.0", "end")
        self.motivo_txt.insert("1.0", self._motivo_ph)
        self.motivo_txt.configure(text_color=("#9CA3AF", "#6b7280"))
        self.motivo_txt.bind("<FocusIn>", self._motivo_focus_in)
        self.motivo_txt.bind("<FocusOut>", self._motivo_focus_out)

    def _motivo_focus_in(self, *_):
        if self._motivo_placeholder_on:
            self.motivo_txt.delete("1.0", "end")
            self.motivo_txt.configure(text_color=("black", "white"))
            self._motivo_placeholder_on = False

    def _motivo_focus_out(self, *_):
        content = (self.motivo_txt.get("1.0", "end").strip())
        if not content:
            self._init_motivo_placeholder()

    def _get_motivo_text(self) -> str:
        if self._motivo_placeholder_on: return ""
        return self.motivo_txt.get("1.0", "end").strip()

    def _set_motivo_text(self, text: str):
        self._motivo_placeholder_on = False
        self.motivo_txt.delete("1.0", "end")
        self.motivo_txt.insert("1.0", text)
        self.motivo_txt.configure(text_color=("black", "white"))

    def _insert_motivo_example(self):
        example = ("Ajusté velocidad de inyección 50->65 mm/s para mejorar llenado en nervaduras; "
                   "espero reducir cortinas y estabilizar peso del disparo (+0.2g).")
        cur = self._get_motivo_text()
        if not cur:
            self._set_motivo_text(example)
        else:
            if not cur.endswith("\n"):
                self.motivo_txt.insert("end", "\n")
            self.motivo_txt.insert("end", example)

    def _mark_dirty(self): self._dirty = True

    def _back_menu_guard(self):
        if self._dirty and not messagebox.askyesno("Cambios sin guardar", "Tienes cambios sin guardar. ¿Descartar?"):
            return
        self._dirty = False
        self._preview_version = None
        self._status_var.set("")
        if hasattr(self.app, "go_menu"):
            self.app.go_menu()

    def _on_pick_mold(self):
        mid = (self.mold_var.get() or "").strip()
        self.current_mold = "" if (not mid or mid.startswith("—")) else mid
        for v in self.vars.values(): v.set("")
        self._dirty = False
        self._preview_version = None
        self._status_var.set("")
        self._init_motivo_placeholder()
        if self.current_mold:
            data = _load_json(self.current_mold)
            for k, v in self.vars.items():
                v.set(str(data.get(k, "")))
            if "date_of_entry" in self.vars and not self.vars["date_of_entry"].get().strip():
                self.vars["date_of_entry"].set(datetime.now().strftime("%Y-%m-%d"))

    def _collect(self) -> dict:
        out = {}
        for k, v in self.vars.items():
            val = (v.get() or "").strip()
            if k in NUM_FIELDS: val = _cast_numeric(val)
            out[k] = val
        return out

    def _resolve_usuario(self) -> str:
        try:
            op = getattr(self.app, "operador", "")
            return op.get() if hasattr(op, "get") else str(op or "")
        except Exception:
            return ""

    def _save(self, *_):
        if not self.current_mold:
            messagebox.showwarning("Molde", "Selecciona un molde."); return
        motivo = self._get_motivo_text()
        if not motivo:
            messagebox.showwarning("Motivo requerido", "Escribe el motivo del cambio para guardar."); return
        if "date_of_entry" in self.vars and not self.vars["date_of_entry"].get().strip():
            self.vars["date_of_entry"].set(datetime.now().strftime("%Y-%m-%d"))

        old = _load_json(self.current_mold)
        new = self._collect()

        diffs = []
        for k, nv in new.items():
            ov = str(old.get(k, ""))
            if (nv or "") != (ov or ""):
                diffs.append(f"{k}: '{ov}' -> '{nv}'")
        if not diffs:
            messagebox.showinfo("Recetas de Máquina", "No hubo cambios en los parámetros.")
            return

        _save_json(self.current_mold, new)

        usuario = self._resolve_usuario()
        diffs_text = " | ".join(diffs)
        ver = _save_version_snapshot(self.current_mold, new, usuario, motivo, diffs_text)

        exists = os.path.exists(HISTORY_CSV)
        with open(HISTORY_CSV, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if not exists:
                w.writerow(["ts", "molde_id", "usuario", "motivo", "cambios", "version"])
            w.writerow([datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        self.current_mold, usuario, motivo, diffs_text, ver])

        self._init_motivo_placeholder()
        self._dirty = False
        self._preview_version = None
        self._status_var.set(f"Guardado como {ver}.")
        messagebox.showinfo("Recetas de Máquina", f"Parámetros guardados.\nSe creó snapshot: {ver}")

    def _clear_all(self):
        if not self.current_mold:
            if messagebox.askyesno("Limpiar", "¿Limpiar todos los campos visibles?"):
                for v in self.vars.values(): v.set("")
            return
        if messagebox.askyesno("Limpiar", f"¿Limpiar todos los campos del molde {self.current_mold}?"):
            for v in self.vars.values(): v.set("")
            self._dirty = True

    # --------------------------- Historial / Versiones ---------------------------
    def _open_history(self):
        if not self.current_mold:
            messagebox.showwarning("Historial", "Selecciona un molde para ver su historial."); return

        top = tk.Toplevel(self); top.title(f"Historial y versiones — {self.current_mold}")
        try: top.state('zoomed')
        except Exception: top.geometry("1280x800+60+40")

        wrap = ctk.CTkFrame(top, fg_color=("white", "#0e1117")); wrap.pack(fill="both", expand=True)

        # Barra superior
        bar = ctk.CTkFrame(wrap, fg_color="transparent"); bar.pack(fill="x", padx=12, pady=(10, 6))
        ctk.CTkLabel(bar, text=f"Molde: {self.current_mold}", font=ctk.CTkFont(size=16, weight="bold")).pack(side="left")
        srch_var = tk.StringVar()
        ctk.CTkEntry(bar, textvariable=srch_var, placeholder_text="Buscar en motivo/cambios...", width=380
                    ).pack(side="left", padx=(12,4))
        ctk.CTkButton(bar, text="Buscar", command=lambda: load_versions()).pack(side="left")
        ctk.CTkButton(bar, text="Mostrar todo", command=lambda: (srch_var.set(""), load_versions())
                     ).pack(side="left", padx=(8,0))
        ctk.CTkButton(bar, text="Abrir carpeta de versiones",
                      command=lambda: self._open_folder(_versions_dir(self.current_mold))
                     ).pack(side="right")
        ctk.CTkButton(bar, text="Cerrar", fg_color="#6b7280", hover_color="#4b5563",
                      command=top.destroy).pack(side="right", padx=(8,0))

        # Paneles
        main = ctk.CTkFrame(wrap, fg_color="transparent"); main.pack(fill="both", expand=True, padx=12, pady=(0,12))
        main.grid_columnconfigure(0, weight=3)
        main.grid_columnconfigure(1, weight=5)
        main.grid_rowconfigure(0, weight=1)

        # IZQUIERDA: versiones
        left = ctk.CTkFrame(main, fg_color=("white", "#111827")); left.grid(row=0, column=0, sticky="nsew", padx=(0,8))
        cols = ("version","fecha","usuario","motivo","cambios")
        tree = ttk.Treeview(left, columns=cols, show="headings", height=24)
        headers = [("version","Versión",80),("fecha","Fecha/Hora",150),("usuario","Usuario",140),
                   ("motivo","Motivo (resumen)",280),("cambios","Cambios (preview)",400)]
        for key, text, w in headers:
            tree.heading(key, text=text); tree.column(key, width=w, anchor="w")
        vsb = ttk.Scrollbar(left, orient="vertical", command=tree.yview); tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="both", expand=True, padx=12, pady=12); vsb.pack(side="left", fill="y", pady=12)

        # DERECHA: detalle (grilla)
        right = ctk.CTkFrame(main, fg_color=("white", "#111827")); right.grid(row=0, column=1, sticky="nsew", padx=(8,0))
        header_box = ctk.CTkFrame(right, fg_color="transparent"); header_box.pack(fill="x", padx=12, pady=(12,0))
        detail_title = ctk.CTkLabel(header_box, text="Detalle de versión", font=ctk.CTkFont(size=14, weight="bold"))
        detail_title.pack(side="left")

        export_box = ctk.CTkFrame(header_box, fg_color="transparent"); export_box.pack(side="right")
        btn_xlsx = ctk.CTkButton(export_box, text="⬇ Exportar Excel", state="disabled")
        btn_pdf  = ctk.CTkButton(export_box, text="⬇ Exportar PDF", state="disabled")
        btn_pdf.pack(side="right", padx=(8,0)); btn_xlsx.pack(side="right")

        detail_wrap = ctk.CTkScrollableFrame(right, fg_color=("white","#111827"))
        detail_wrap.pack(fill="both", expand=True, padx=12, pady=(6,12))

        # placeholder de secciones
        def clear_detail():
            for w in detail_wrap.winfo_children():
                w.destroy()

        # ----- grilla estilo Excel por secciones -----
        def section(title):
            card = ctk.CTkFrame(detail_wrap, corner_radius=12, fg_color=("white", "#111a27"))
            card.pack(fill="x", padx=6, pady=6)
            ctk.CTkLabel(card, text=title, font=ctk.CTkFont("Helvetica", 12, "bold")).pack(anchor="w", padx=10, pady=(10,6))
            grid = ctk.CTkFrame(card, fg_color="transparent"); grid.pack(fill="x", padx=10, pady=(0,10))
            return grid

        def row_labeled(grid, r, label, values, start_col=0):
            ctk.CTkLabel(grid, text=label).grid(row=r, column=start_col, sticky="w", padx=4, pady=4)
            for i, val in enumerate(values, start=1):
                e = ctk.CTkEntry(grid, width=110, justify="center")
                e.insert(0, str(val))
                e.configure(state="disabled")
                e.grid(row=r, column=start_col+i, sticky="ew", padx=4, pady=4)

        # datos actuales en detalle
        last_snap = {"_meta": {}}

        def load_versions():
            tree.delete(*tree.get_children())
            q = (srch_var.get() or "").lower()
            snaps = _list_versions(self.current_mold)
            if not snaps:
                current = _load_json(self.current_mold)
                if current:
                    usuario = self._resolve_usuario()
                    motivo = "(baseline) Estado inicial importado desde receta actual"
                    _save_version_snapshot(self.current_mold, current, usuario, motivo, "")
                    snaps = _list_versions(self.current_mold)
            for ver, path in snaps:
                try:
                    with open(path, "r", encoding="utf-8") as f:
                        snap = json.load(f)
                except Exception:
                    continue
                meta = snap.get("_meta", {})
                ts = meta.get("saved_at",""); usuario = meta.get("usuario","")
                motivo = meta.get("motivo",""); cambios = meta.get("diffs","")
                rowtxt = (motivo + " " + cambios).lower()
                if q and q not in rowtxt: continue
                tree.insert("", "end", values=(ver, ts, usuario, motivo, cambios))

            # --- AUTOSELECT last ---
            items = tree.get_children()
            if items:
                tree.selection_set(items[-1]); tree.see(items[-1]); on_select()

        def on_select(*_):
            nonlocal last_snap
            clear_detail()
            sel = tree.selection()
            if not sel:
                detail_title.configure(text="Detalle de versión")
                btn_xlsx.configure(state="disabled"); btn_pdf.configure(state="disabled")
                return
            ver = tree.item(sel[0], "values")[0]
            snap = _load_version_snapshot(self.current_mold, ver)
            last_snap = snap
            meta = snap.get("_meta", {})
            clean = dict(snap); clean.pop("_meta", None)

            detail_title.configure(text=f"Detalle de versión — {ver}")
            btn_xlsx.configure(state="normal")
            btn_pdf.configure(state="normal")

            # Parameter overview
            g = section("Parameter overview")
            g.grid_columnconfigure((0,1,2,3,4,5), weight=1)
            row_labeled(g, 0, "Program",      [clean.get("program",""), "Date of entry", clean.get("date_of_entry",""),
                                               "Cavities", clean.get("cavities","")], start_col=0)
            row_labeled(g, 1, "Mould desig.", [clean.get("mould_desig",""), "Machine", clean.get("machine",""),
                                               "Material", clean.get("material","")], start_col=0)

            # Key data
            k = section("Key data")
            items = [
                ("Cycle time [s]", "cycle_time_s"),
                ("Injection time [s]", "injection_time_s"),
                ("Holding press. time [s]", "holding_press_time_s"),
                ("Rem. cooling time [s]", "rem_cooling_time_s"),
                ("Dosage time [s]", "dosage_time_s"),
                ("Screw stroke [mm]", "screw_stroke_mm"),
                ("Mould stroke [mm]", "mould_stroke_mm"),
                ("Ejector stroke [mm]", "ejector_stroke_mm"),
                ("Shot weight [g]", "shot_weight_g"),
                ("Plasticising flow [kg/h]", "plasticising_flow_kgh"),
                ("Dosage capacity [g/s]", "dosage_capacity_gs"),
                ("Dosage volume [ccm]", "dosage_volume_ccm"),
                ("Material cushion [ccm]", "material_cushion_ccm"),
                ("max. inj. pressure [bar]", "max_inj_pressure_bar"),
            ]
            k.grid_columnconfigure((0,1,2), weight=1)
            rr = 0
            for label, fid in items:
                row_labeled(k, rr, label, [clean.get(fid,"")], start_col=0)
                rr += 1

            # Injection unit
            iu = section("Injection unit")
            iu.grid_columnconfigure((0,1,2,3), weight=1)
            row_labeled(iu, 0, "Screw Ø [mm]", [clean.get("screw_d_mm",""), "Pcs. 1", clean.get("pcs_1","")], start_col=0)

            # Injection
            inj = section("Injection")
            inj.grid_columnconfigure((0,1,2,3,4), weight=1)
            row_labeled(inj, 0, "Injection press. limiting [bar]",
                        [clean.get("inj_press_lim_1",""), clean.get("inj_press_lim_2",""), clean.get("inj_press_lim_3","")])
            row_labeled(inj, 1, "Injection speed [mm/s]",
                        [clean.get("inj_speed_1",""), clean.get("inj_speed_2",""), clean.get("inj_speed_3","")])
            row_labeled(inj, 2, "End of stage [mm]",
                        [clean.get("inj_end_stage_mm_1",""), clean.get("inj_end_stage_mm_2",""), clean.get("inj_end_stage_mm_3","")])
            row_labeled(inj, 3, "Injection flow [ccm/s]",
                        [clean.get("inj_flow_1",""), clean.get("inj_flow_2",""), clean.get("inj_flow_3","")])
            row_labeled(inj, 4, "End of stage [ccm]",
                        [clean.get("inj_end_stage_ccm_1",""), clean.get("inj_end_stage_ccm_2",""), clean.get("inj_end_stage_ccm_3","")])

            # Plasticizing
            pl = section("Plasticizing (St.1)")
            pl.grid_columnconfigure((0,1), weight=1)
            row_labeled(pl, 0, "Screw speed [m/min]", [clean.get("plast_screw_speed","")])
            row_labeled(pl, 1, "Back pressure [bar]", [clean.get("plast_back_pressure","")])
            row_labeled(pl, 2, "End of stage [ccm]", [clean.get("plast_end_stage_ccm","")])

            # Holding pressure
            hp = section("Holding pressure (Pcs.2)")
            hp.grid_columnconfigure((0,1,2,3,4), weight=1)
            row_labeled(hp, 0, "Time [s]",     [clean.get("hp_time_1",""), clean.get("hp_time_2",""), clean.get("hp_time_3","")])
            row_labeled(hp, 1, "Pressure [bar]", [clean.get("hp_press_1",""), clean.get("hp_press_2",""),
                                                  clean.get("hp_press_3",""), clean.get("hp_press_4","")])

            # Temperatures
            tp = section("Temperatures (1..5)")
            tp.grid_columnconfigure((0,1,2,3,4,5), weight=1)
            row_labeled(tp, 0, "Cylinder temp. [°C]", [clean.get("temp_c1",""), clean.get("temp_c2",""),
                                                       clean.get("temp_c3",""), clean.get("temp_c4",""),
                                                       clean.get("temp_c5","")])
            row_labeled(tp, 1, "Tolerances [°C]", [clean.get("tol_c1",""), clean.get("tol_c2",""),
                                                   clean.get("tol_c3",""), clean.get("tol_c4",""),
                                                   clean.get("tol_c5","")])
            row_labeled(tp, 2, "Feed yoke temperature [°C]", [clean.get("feed_yoke_temp","")])
            row_labeled(tp, 3, "Lower enable tol. [°C]",     [clean.get("lower_enable_tol","")])
            row_labeled(tp, 4, "Upper switch-off tol. [°C]", [clean.get("upper_switch_off_tol","")])

            # Mould movements Opening
            mo = section("Mould movements — Opening (St.1 / St.2 / St.3)")
            mo.grid_columnconfigure((0,1,2,3), weight=1)
            row_labeled(mo, 0, "End of stage [mm]", [clean.get("open_end_mm_1",""), clean.get("open_end_mm_2",""), clean.get("open_end_mm_3","")])
            row_labeled(mo, 1, "Speed [mm/s]",      [clean.get("open_speed_1",""), clean.get("open_speed_2",""), clean.get("open_speed_3","")])
            row_labeled(mo, 2, "Force [kN]",        [clean.get("open_force_1",""), clean.get("open_force_2",""), clean.get("open_force_3","")])

            # Mould movements Closing
            mc = section("Mould movements — Closing (St.1 / St.2 / St.3 / An. HD)")
            mc.grid_columnconfigure((0,1,2,3,4), weight=1)
            row_labeled(mc, 0, "End of stage [mm]", [clean.get("close_end_mm_1",""), clean.get("close_end_mm_2",""),
                                                     clean.get("close_end_mm_3",""), clean.get("close_end_mm_4","")])
            row_labeled(mc, 1, "Speed [mm/s]",      [clean.get("close_speed_1",""), clean.get("close_speed_2",""),
                                                     clean.get("close_speed_3",""), clean.get("close_speed_4","")])
            row_labeled(mc, 2, "Force [kN]",        [clean.get("close_force_1",""), clean.get("close_force_2",""),
                                                     clean.get("close_force_3","")])

            # Clamping
            cl = section("Clamping")
            cl.grid_columnconfigure((0,1), weight=1)
            row_labeled(cl, 0, "Mould closed [kN]", [clean.get("mould_closed_kn","")])

        # ---------------------- EXPORTS (integrados) ----------------------
        def export_excel():
            sel = tree.selection()
            if not sel:
                return
            ver = tree.item(sel[0], "values")[0]
            ver_label = f"V.{ver[1:]}" if str(ver).lower().startswith("v") else str(ver)

            # 1) Elegir dónde guardar
            default_name = f"{self.current_mold}_{ver}.xlsx"
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialfile=default_name
            )
            if not path:
                return

            try:
                # 2) Encontrar plantilla
                template_path = _find_excel_template()
                if not template_path:
                    # Si no está, deja elegirla manualmente una sola vez
                    messagebox.showwarning(
                        "Plantilla no encontrada",
                        "No encontré 'formatorecta.xlsx' en las rutas conocidas.\n"
                        "Selecciona la plantilla (se usará esta vez):"
                    )
                    template_path = filedialog.askopenfilename(
                        title="Selecciona formatorecta.xlsx",
                        filetypes=[("Excel", "*.xlsx")]
                    )
                    if not template_path:
                        return

                # 3) Obtener snapshot del árbol (la versión seleccionada)
                snap = _load_version_snapshot(self.current_mold, ver)
                if not snap:
                    messagebox.showwarning("Exportar Excel", "No se pudo cargar el snapshot de la versión.")
                    return
                # Limpia metadatos para que solo queden campos
                snap_clean = dict(snap)
                snap_clean.pop("_meta", None)

                # 4) Exportar a plantilla usando el EXCEL_MAP
                #    Si tu plantilla usa una hoja específica, pásala por 'sheet_name="NombreHoja"'
                _export_snapshot_to_template(snapshot=snap_clean, out_path=path, template_path=template_path)

                messagebox.showinfo("Exportar Excel", f"Archivo generado correctamente:\n{path}")
            except ModuleNotFoundError:
                messagebox.showwarning("Dependencia faltante",
                    "Falta openpyxl para exportar en plantilla:\n\npip install openpyxl")
            except Exception as e:
                messagebox.showerror("Exportar Excel", f"No se pudo exportar:\n{e}")

        def export_pdf():
            sel = tree.selection()
            if not sel:
                return
            ver = tree.item(sel[0], "values")[0]
            ver_label = f"V.{ver[1:]}" if str(ver).lower().startswith("v") else str(ver)
            title_txt = f"Receta {ver_label}"
            path = filedialog.asksaveasfilename(defaultextension=".pdf",
                                                filetypes=[("PDF", "*.pdf")],
                                                initialfile=f"{self.current_mold}_{ver}.pdf")
            if not path:
                return
            try:
                from reportlab.lib.pagesizes import A4, landscape
                from reportlab.pdfgen import canvas
                from reportlab.lib.units import mm
                from reportlab.lib.utils import ImageReader

                snap = dict(last_snap); _ = snap.pop("_meta", {})
                W, H = landscape(A4)
                margin = 12*mm
                col_gap = 18*mm
                col_w = (W - margin*2 - col_gap) / 2.0
                row_h = 10*mm

                logo_path = LOGO_PATH if os.path.exists(LOGO_PATH) else ""
                c = canvas.Canvas(path, pagesize=landscape(A4))
                c.setStrokeColorRGB(0.7, 0.7, 0.7)
                c.setLineWidth(0.5)

                def new_page():
                    c.showPage()
                    return W, H, margin, H - margin

                def draw_head(txt):
                    c.setFillColorRGB(0.86, 0.92, 1)
                    c.rect(x, y - row_h, col_w, row_h, stroke=0, fill=1)
                    c.setFillColorRGB(0.11, 0.30, 0.85)
                    c.setFont("Helvetica-Bold", 12)
                    c.drawString(x + 2, y - row_h + 3, _safe_pdf(txt))
                    c.setFillColorRGB(0, 0, 0)

                def draw_box(lbl, val, w, h=row_h, center=False):
                    c.setFillColorRGB(0.96, 0.96, 0.96)
                    c.rect(x, y - h, w, h, stroke=1, fill=1)
                    c.setFillColorRGB(0, 0, 0)
                    c.setFont("Helvetica-Bold", 9)
                    c.drawString(x+2, y - 12, _safe_pdf(lbl))
                    if val is not None:
                        c.setFont("Helvetica", 10)
                        if center:
                            c.drawCentredString(x + w/2, y - h/2 - 3, _safe_pdf(val))
                        else:
                            c.drawString(x+2, y - h + 4, _safe_pdf(val))

                def advance(h=row_h+6):
                    nonlocal y, x, col
                    y -= h
                    if y < margin + row_h*5:
                        if col == 0:
                            col = 1; x = margin + col_w + col_gap; y = H - margin - 24
                        else:
                            col = 0; x = margin
                            _, _, _, y0 = new_page(); y = y0 - 24
                        # logo + título por página
                        if logo_path:
                            try:
                                img = ImageReader(logo_path)
                                c.drawImage(img, W-60*mm, H-18*mm, width=50*mm, height=12*mm,
                                            preserveAspectRatio=True, mask='auto')
                            except Exception:
                                pass
                        c.setFont("Helvetica-Bold", 14)
                        c.drawString(margin, H - margin - 8, _safe_pdf(title_txt))

                # logo + título primera página
                if logo_path:
                    try:
                        img = ImageReader(logo_path)
                        c.drawImage(img, W-60*mm, H-18*mm, width=50*mm, height=12*mm,
                                    preserveAspectRatio=True, mask='auto')
                    except Exception:
                        pass
                c.setFont("Helvetica-Bold", 14)
                c.drawString(margin, H - margin - 8, _safe_pdf(title_txt))

                x, y = margin, H - margin - 24
                col = 0

                # Cabecera
                draw_box("Program", snap.get("program",""), w=40*mm); advance()
                draw_box("Date of entry:", snap.get("date_of_entry",""), w=60*mm); advance()
                draw_box("Cavities", snap.get("cavities",""), w=30*mm); advance()
                draw_box("Mould desig.", snap.get("mould_desig",""), w=40*mm); advance()
                draw_box("Machine", snap.get("machine",""), w=40*mm); advance()
                draw_box("Material", snap.get("material",""), w=40*mm); advance(h=10)

                # Injection unit
                draw_head("Injection unit"); advance(h=12)
                draw_box("Screw Ø [mm]", snap.get("screw_d_mm",""), w=35*mm); advance()
                draw_box("Pcs. 1", snap.get("pcs_1",""), w=25*mm); advance(h=14)

                # Forzar salto para Key data si falta espacio
                if y < margin + row_h*18 and col == 0:
                    col = 1; x = margin + col_w + col_gap; y = H - margin - 24

                # Key data
                draw_head("Key data"); advance(h=12)
                kd = [
                    ("Cycle time [s]", "cycle_time_s"),
                    ("Injection time [s]", "injection_time_s"),
                    ("Holding press. time [s]", "holding_press_time_s"),
                    ("Rem. cooling time [s]", "rem_cooling_time_s"),
                    ("Dosage time [s]", "dosage_time_s"),
                    ("Screw stroke [mm]", "screw_stroke_mm"),
                    ("Mould stroke [mm]", "mould_stroke_mm"),
                    ("Ejector stroke [mm]", "ejector_stroke_mm"),
                    ("Shot weight [g]", "shot_weight_g"),
                    ("Plasticising flow [kg/h]", "plasticising_flow_kgh"),
                    ("Dosage capacity [g/s]", "dosage_capacity_gs"),
                    ("Dosage volume [ccm]", "dosage_volume_ccm"),
                    ("Material cushion [ccm]", "material_cushion_ccm"),
                    ("max. inj. pressure [bar]", "max_inj_pressure_bar"),
                ]
                for lab, fid in kd:
                    draw_box(lab, snap.get(fid,""), w=col_w); advance()

                # Volver a primera col si conviene
                col = 0; x = margin; y = y if y > H/2 else H - margin - 24

                # Injection (3 columnas)
                draw_head("Injection"); advance(h=12)
                def triple(lbl, k1, k2, k3):
                    wlbl = 65*mm; wcell = (col_w - wlbl)/3.0
                    draw_box(lbl, None, w=wlbl)
                    c.rect(x+wlbl, y - row_h, wcell, row_h, 1, 0)
                    c.drawCentredString(x+wlbl + wcell/2, y - row_h/2 - 3, _safe_pdf(snap.get(k1,"")))
                    c.rect(x+wlbl+wcell, y - row_h, wcell, row_h, 1, 0)
                    c.drawCentredString(x+wlbl + wcell*1.5, y - row_h/2 - 3, _safe_pdf(snap.get(k2,"")))
                    c.rect(x+wlbl+wcell*2, y - row_h, wcell, row_h, 1, 0)
                    c.drawCentredString(x+wlbl + wcell*2.5, y - row_h/2 - 3, _safe_pdf(snap.get(k3,"")))
                    advance()
                triple("Injection press. limiting [bar]", "inj_press_lim_1","inj_press_lim_2","inj_press_lim_3")
                triple("Injection speed [mm/s]",          "inj_speed_1","inj_speed_2","inj_speed_3")
                triple("End of stage [mm]",               "inj_end_stage_mm_1","inj_end_stage_mm_2","inj_end_stage_mm_3")
                triple("Injection flow [ccm/s]",          "inj_flow_1","inj_flow_2","inj_flow_3")
                triple("End of stage [ccm]",              "inj_end_stage_ccm_1","inj_end_stage_ccm_2","inj_end_stage_ccm_3")

                # Plasticizing
                draw_head("Plasticizing (St.1)"); advance(h=12)
                draw_box("Screw speed [m/min]", snap.get("plast_screw_speed",""), w=col_w); advance()
                draw_box("Back pressure [bar]", snap.get("plast_back_pressure",""), w=col_w); advance()
                draw_box("End of stage [ccm]",  snap.get("plast_end_stage_ccm",""), w=col_w); advance(h=10)

                # Holding pressure
                draw_head("Holding pressure (Pcs.2)"); advance(h=12)
                triple("Time [s]", "hp_time_1","hp_time_2","hp_time_3")
                # presión (4 columnas)
                wlbl = 65*mm; wcell = (col_w - wlbl)/4.0
                draw_box("Pressure [bar]", None, w=wlbl)
                keys = ["hp_press_1","hp_press_2","hp_press_3","hp_press_4"]
                for i,k in enumerate(keys):
                    cx = x + wlbl + i*wcell
                    c.rect(cx, y - row_h, wcell, row_h, 1, 0)
                    c.drawCentredString(cx + wcell/2, y - row_h/2 - 3, _safe_pdf(snap.get(k,"")))
                advance(h=row_h+4)

                # Temperatures
                draw_head("Temperatures"); advance(h=12)
                def five(lbl, keys):
                    wlbl = 65*mm; wcell = (col_w - wlbl)/5.0
                    draw_box(lbl, None, w=wlbl)
                    for i,k in enumerate(keys):
                        cx = x + wlbl + i*wcell
                        c.rect(cx, y - row_h, wcell, row_h, 1, 0)
                        c.drawCentredString(cx + wcell/2, y - row_h/2 - 3, _safe_pdf(snap.get(k,"")))
                    advance()
                five("Cylinder temp. [°C]", ["temp_c1","temp_c2","temp_c3","temp_c4","temp_c5"])
                five("Tolerances [°C]",     ["tol_c1","tol_c2","tol_c3","tol_c4","tol_c5"])
                draw_box("Feed yoke temperature [°C]", snap.get("feed_yoke_temp",""), w=col_w); advance()
                # doble columna en fila
                whalf = (col_w - 4*mm)/2
                draw_box("Lower enable tol. [°C]", snap.get("lower_enable_tol",""), w=whalf)
                c.rect(x+whalf+4*mm, y - row_h, whalf, row_h, 1, 0)
                c.setFont("Helvetica-Bold", 9); c.drawString(x+whalf+4*mm+2, y - 12, _safe_pdf("Upper switch-off tol. [°C]"))
                c.setFont("Helvetica", 10); c.drawString(x+whalf+4*mm+2, y - row_h + 4, _safe_pdf(snap.get("upper_switch_off_tol","")))
                advance(h=row_h+6)

                # Opening
                draw_head("Mould movements — Opening (St.1 / St.2 / St.3)"); advance(h=12)
                triple("End of stage [mm]", "open_end_mm_1","open_end_mm_2","open_end_mm_3")
                triple("Speed [mm/s]",      "open_speed_1","open_speed_2","open_speed_3")
                triple("Force [kN]",        "open_force_1","open_force_2","open_force_3")

                # Closing
                draw_head("Mould movements — Closing (St.1 / St.2 / St.3 / An. HD)"); advance(h=12)
                def quad(lbl, keys):
                    wlbl = 65*mm; wcell = (col_w - wlbl)/4.0
                    draw_box(lbl, None, w=wlbl)
                    for i,k in enumerate(keys):
                        cx = x + wlbl + i*wcell
                        c.rect(cx, y - row_h, wcell, row_h, 1, 0)
                        c.drawCentredString(cx + wcell/2, y - row_h/2 - 3, _safe_pdf(snap.get(k,"")))
                    advance()
                quad("End of stage [mm]", ["close_end_mm_1","close_end_mm_2","close_end_mm_3","close_end_mm_4"])
                quad("Speed [mm/s]",      ["close_speed_1","close_speed_2","close_speed_3","close_speed_4"])
                # 3 columnas para force
                wlbl = 65*mm; wcell = (col_w - wlbl)/4.0
                draw_box("Force [kN]", None, w=wlbl)
                for i,k in enumerate(["close_force_1","close_force_2","close_force_3"]):
                    cx = x + wlbl + i*wcell
                    c.rect(cx, y - row_h, wcell, row_h, 1, 0)
                    c.drawCentredString(cx + wcell/2, y - row_h/2 - 3, _safe_pdf(snap.get(k,"")))
                advance()

                # Clamping
                draw_head("Clamping"); advance(h=12)
                draw_box("Mould closed [kN]", snap.get("mould_closed_kn",""), w=col_w); advance()

                c.showPage(); c.save()
                messagebox.showinfo("Exportar PDF","Archivo PDF generado correctamente.")
            except ModuleNotFoundError:
                messagebox.showwarning("Dependencia faltante",
                    "Para exportar PDF instala reportlab:\n\npip install reportlab")
            except Exception as e:
                messagebox.showerror("Exportar PDF", f"No se pudo exportar:\n{e}")

        tree.bind("<<TreeviewSelect>>", on_select)
        btn_xlsx.configure(command=export_excel)
        btn_pdf.configure(command=export_pdf)
        load_versions()

    # --------------------------- Importar Excel simple ---------------------------
    def _import_excel(self):
        try:
            from openpyxl import load_workbook
        except Exception:
            messagebox.showwarning("Falta dependencia", "Instala openpyxl: pip install openpyxl")
            return
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx;*.xlsm;*.xls")])
        if not path: return
        try:
            wb = load_workbook(path, data_only=True, keep_vba=True); ws = wb.active
            hits, misses = 0, 0
            for fid, cell in EXCEL_MAP.items():
                try: val = ws[cell].value
                except Exception: val = None
                if fid in self.vars and val is not None:
                    self.vars[fid].set(_cast_numeric(val) if fid in NUM_FIELDS else str(val))
                    hits += 1
                else:
                    misses += 1
            self._dirty = True
            messagebox.showinfo("Importar",
                                f"Se mapearon {hits} campo(s). Revisa EXCEL_MAP para cubrir más campos. "
                                f"Sin valor: {misses}.")
        except Exception as e:
            messagebox.showerror("Importar Excel", f"No se pudo leer el archivo:\n{e}")

    # --------------------------- Utilidades ---------------------------
    def _open_folder(self, path):
        try:
            if os.path.isdir(path):
                if sys.platform.startswith("win"):
                    os.startfile(path)
                elif sys.platform == "darwin":
                    os.system(f'open "{path}"')
                else:
                    os.system(f'xdg-open "{path}"')
            else:
                folder = os.path.dirname(path)
                if sys.platform.startswith("win"):
                    os.startfile(folder)
                elif sys.platform == "darwin":
                    os.system(f'open "{folder}"')
                else:
                    os.system(f'xdg-open "{folder}"')
        except Exception as e:
            messagebox.showwarning("Abrir carpeta", f"No se pudo abrir:\n{e}")
