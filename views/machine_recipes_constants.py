# -*- coding: utf-8 -*-
# Recetas de Máquina (layout estilo SELOGICA) — versiones + historial con grilla y export
from .base import BASE_DIR
import os, json, csv, sys, re
from datetime import datetime
from openpyxl.utils import range_boundaries, get_column_letter, column_index_from_string
from openpyxl.cell.cell import MergedCell

# --- Detectar Excel COM una sola vez ---
try:
    import win32com.client as win32  # pywin32
    HAS_EXCEL_COM = True
except Exception:
    HAS_EXCEL_COM = False

# ---------- Paths ----------
RECIPES_DIR = os.path.join(BASE_DIR, "machine_recipes")
HISTORY_CSV = os.path.join(RECIPES_DIR, "_history.csv")
os.makedirs(RECIPES_DIR, exist_ok=True)

DATA_SCHEMA_VERSION = 2

# ---------- Excel helpers ----------
EXCEL_MAP = {
    # ---------- Encabezado ----------
    "program": "DEFG:4",
    "mould_desig": "DEFG:5",
    "material": "DEFG:6",
    "date_of_entry": "KLMNO:4",
    "cavities": "KLMNO:5",
    "machine": "KLMNO:6",

    # ---------- Injection unit / Key data (superior) ----------
    "screw_diameter_mm": "J:9",

    # Key data (N–O)
    "cycle_time_s": "NO:8",
    "injection_time_s": "NO:9",
    "holding_press_time_s": "NO:10",
    "rem_cooling_time_s": "NO:11",
    "dosage_time_s": "NO:12",
    "screw_stroke_mm": "NO:13",
    "mould_stroke_mm": "NO:14",
    "ejector_stroke_mm": "NO:15",
    "shot_weight_g": "NO:16",
    "plasticising_flow_kg_h": "NO:17",
    "dosage_capacity_g_s": "NO:18",
    "dosage_volume_ccm": "NO:19",
    "material_cushion_ccm": "NO:20",
    "max_injection_pressure_bar": "NO:21",

    # ---------- Injection (filas 11–15) ----------
    "inj_press_limit_bar_1": "I:11",
    "inj_press_limit_bar_2": "H:11",
    "inj_press_limit_bar_3": "G:11",

    "injection_speed_1": "I:12",
    "injection_speed_2": "H:12",
    "injection_speed_3": "G:12",

    "end_of_stage_mm_1": "I:13",
    "end_of_stage_mm_2": "H:13",
    "end_of_stage_mm_3": "G:13",

    "injection_flow_ccm_s_1": "I:14",
    "injection_flow_ccm_s_2": "H:14",
    "injection_flow_ccm_s_3": "G:14",

    "end_of_stage_ccm_1": "I:15",
    "end_of_stage_ccm_2": "H:15",
    "end_of_stage_ccm_3": "G:15",

    # ---------- Plasticizing (filas 18–20) ----------
    "screw_speed_m_min_1": "E:18",
    "back_pressure_bar": "E:19",
    "plasticizing_end_stage_ccm": "E:20",

    # ---------- Holding pressure (filas 23–24) ----------
    "hold_time_s_1": "F:23",
    "hold_time_s_2": "G:23",
    "hold_time_s_3": "H:23",

    "hold_pressure_bar_1": "F:24",
    "hold_pressure_bar_2": "G:24",
    "hold_pressure_bar_3": "H:24",

    # ---------- Temperatures (filas 27–29) ----------
    "cylinder_temp_c_1": "D:27",
    "cylinder_temp_c_2": "E:27",
    "cylinder_temp_c_3": "F:27",
    "cylinder_temp_c_4": "G:27",
    "cylinder_temp_c_5": "H:27",

    "tolerance_c_1": "D:28",
    "tolerance_c_2": "E:28",
    "tolerance_c_3": "F:28",
    "tolerance_c_4": "G:28",
    "tolerance_c_5": "H:28",

    "feed_yoke_temp_c": "D:29",
    "lower_enable_tol_c": "J:30",
    "upper_switch_off_tol_c": "M:31",

    # ---------- Mould movements (31–33) ----------
    # Opening
    "opening_end_stage_mm_1": "E:31",
    "opening_end_stage_mm_2": "F:31",
    "opening_end_stage_mm_3": "G:31",
    "opening_end_stage_mm_4": "H:31",

    "opening_speed_mm_s_1": "E:32",
    "opening_speed_mm_s_2": "F:32",
    "opening_speed_mm_s_3": "G:32",
    "opening_speed_mm_s_4": "H:32",

    "opening_force_kN_1": "E:33",
    "opening_force_kN_2": "F:33",
    "opening_force_kN_3": "G:33",
    "opening_force_kN_4": "H:33",

    # Closing
    "closing_end_stage_mm_1": "K:31",
    "closing_end_stage_mm_2": "L:31",
    "closing_end_stage_mm_3": "M:31",
    "closing_end_stage_mm_4": "N:31",

    "closing_speed_mm_s_1": "K:32",
    "closing_speed_mm_s_2": "L:32",
    "closing_speed_mm_s_3": "M:32",
    "closing_speed_mm_s_4": "N:32",

    "closing_force_kN_1": "K:33",
    "closing_force_kN_2": "L:33",
    "closing_force_kN_3": "M:33",
    "closing_force_kN_4": "N:33",

    # ---------- Clamping ----------
    "mould_closed_kN": "D:37",
}

# (B) Traducción Excel ↔ UI (alias de claves)

ALIAS_EXCEL_TO_UI = {
    # Encabezado
    "program": "program",
    "mould_desig": "mould_desig",
    "material": "material",
    "date_of_entry": "date_of_entry",
    "cavities": "cavities",
    "machine": "machine",

    # Injection unit / key data
    "screw_diameter_mm": "screw_d_mm",

    "cycle_time_s": "cycle_time_s",
    "injection_time_s": "injection_time_s",
    "holding_press_time_s": "holding_press_time_s",
    "rem_cooling_time_s": "rem_cooling_time_s",
    "dosage_time_s": "dosage_time_s",
    "screw_stroke_mm": "screw_stroke_mm",
    "mould_stroke_mm": "mould_stroke_mm",
    "ejector_stroke_mm": "ejector_stroke_mm",
    "shot_weight_g": "shot_weight_g",
    "plasticising_flow_kg_h": "plasticising_flow_kgh",
    "dosage_capacity_g_s": "dosage_capacity_gs",
    "dosage_volume_ccm": "dosage_volume_ccm",
    "material_cushion_ccm": "material_cushion_ccm",
    "max_injection_pressure_bar": "max_inj_pressure_bar",

    # Injection
    "inj_press_limit_bar_1": "inj_press_lim_1",
    "inj_press_limit_bar_2": "inj_press_lim_2",
    "inj_press_limit_bar_3": "inj_press_lim_3",

    "injection_speed_1": "inj_speed_1",
    "injection_speed_2": "inj_speed_2",
    "injection_speed_3": "inj_speed_3",

    "end_of_stage_mm_1": "inj_end_stage_mm_1",
    "end_of_stage_mm_2": "inj_end_stage_mm_2",
    "end_of_stage_mm_3": "inj_end_stage_mm_3",

    "injection_flow_ccm_s_1": "inj_flow_1",
    "injection_flow_ccm_s_2": "inj_flow_2",
    "injection_flow_ccm_s_3": "inj_flow_3",

    "end_of_stage_ccm_1": "inj_end_stage_ccm_1",
    "end_of_stage_ccm_2": "inj_end_stage_ccm_2",
    "end_of_stage_ccm_3": "inj_end_stage_ccm_3",

    # Plasticizing
    "screw_speed_m_min_1": "plast_screw_speed",
    "back_pressure_bar": "plast_back_pressure",
    "plasticizing_end_stage_ccm": "plast_end_stage_ccm",

    # Holding pressure
    "hold_time_s_1": "hp_time_1",
    "hold_time_s_2": "hp_time_2",
    "hold_time_s_3": "hp_time_3",

    "hold_pressure_bar_1": "hp_press_1",
    "hold_pressure_bar_2": "hp_press_2",
    "hold_pressure_bar_3": "hp_press_3",

    # Temperatures
    "cylinder_temp_c_1": "temp_c1",
    "cylinder_temp_c_2": "temp_c2",
    "cylinder_temp_c_3": "temp_c3",
    "cylinder_temp_c_4": "temp_c4",
    "cylinder_temp_c_5": "temp_c5",

    "tolerance_c_1": "tol_c1",
    "tolerance_c_2": "tol_c2",
    "tolerance_c_3": "tol_c3",
    "tolerance_c_4": "tol_c4",
    "tolerance_c_5": "tol_c5",

    "feed_yoke_temp_c": "feed_yoke_temp",
    "lower_enable_tol_c": "lower_enable_tol",
    "upper_switch_off_tol_c": "upper_switch_off_tol",

    # Movements Opening
    "opening_end_stage_mm_1": "open_end_mm_1",
    "opening_end_stage_mm_2": "open_end_mm_2",
    "opening_end_stage_mm_3": "open_end_mm_3",
    "opening_end_stage_mm_4": "open_end_mm_4",

    "opening_speed_mm_s_1": "open_speed_1",
    "opening_speed_mm_s_2": "open_speed_2",
    "opening_speed_mm_s_3": "open_speed_3",
    "opening_speed_mm_s_4": "open_speed_4",

    "opening_force_kN_1": "open_force_1",
    "opening_force_kN_2": "open_force_2",
    "opening_force_kN_3": "open_force_3",
    "opening_force_kN_4": "open_force_4",

    # Movements Closing
    "closing_end_stage_mm_1": "close_end_mm_1",
    "closing_end_stage_mm_2": "close_end_mm_2",
    "closing_end_stage_mm_3": "close_end_mm_3",
    "closing_end_stage_mm_4": "close_end_mm_4",

    "closing_speed_mm_s_1": "close_speed_1",
    "closing_speed_mm_s_2": "close_speed_2",
    "closing_speed_mm_s_3": "close_speed_3",
    "closing_speed_mm_s_4": "close_speed_4",

    "closing_force_kN_1": "close_force_1",
    "closing_force_kN_2": "close_force_2",
    "closing_force_kN_3": "close_force_3",
    "closing_force_kN_4": "close_force_4",

    # Clamping
    "mould_closed_kN": "mould_closed_kn",
}

NUM_FIELDS = {
    "cycle_time_s","injection_time_s","holding_press_time_s","rem_cooling_time_s",
    "dosage_time_s","screw_stroke_mm","mould_stroke_mm","ejector_stroke_mm",
    "shot_weight_g","plasticising_flow_kgh","dosage_capacity_gs","dosage_volume_ccm",
    "material_cushion_ccm","max_inj_pressure_bar","screw_d_mm","pcs_1",
    "inj_press_lim_1","inj_press_lim_2","inj_press_lim_3",
    "inj_speed_1","inj_speed_2","inj_speed_3",
    "inj_end_stage_mm_1","inj_end_stage_mm_2","inj_end_stage_mm_3",
    "inj_flow_1","inj_flow_2","inj_flow_3",
    "inj_end_stage_ccm_1","inj_end_stage_ccm_2","inj_end_stage_ccm_3",
    "plast_screw_speed","plast_back_pressure","plast_end_stage_ccm",
    "hp_time_1","hp_time_2","hp_time_3",
    "hp_press_1","hp_press_2","hp_press_3","hp_press_4",
    "temp_c1","temp_c2","temp_c3","temp_c4","temp_c5",
    "tol_c1","tol_c2","tol_c3","tol_c4","tol_c5",
    "feed_yoke_temp","lower_enable_tol","upper_switch_off_tol",
    "open_end_mm_1","open_end_mm_2","open_end_mm_3","open_end_mm_4",
    "open_speed_1","open_speed_2","open_speed_3","open_speed_4",
    "open_force_1","open_force_2","open_force_3","open_force_4",
    "close_end_mm_1","close_end_mm_2","close_end_mm_3","close_end_mm_4",
    "close_speed_1","close_speed_2","close_speed_3","close_speed_4",
    "close_force_1","close_force_2","close_force_3","close_force_4",
    "mould_closed_kn",
}

# === Usa tu plantilla predefinida formatorecta.xlsx ===
# Intenta estas rutas en orden; usa la primera que exista.
TEMPLATE_CANDIDATES = [
    os.path.join(BASE_DIR, "formatorecta.xlsx"),
    os.path.join(RECIPES_DIR, "formatorecta.xlsx"),
    os.path.join(BASE_DIR, "assets", "formatorecta.xlsx"),
    "/mnt/data/formatorecta.xlsx",  # por si ya lo tienes ahí
]

def _find_excel_template():
    for p in TEMPLATE_CANDIDATES:
        if os.path.exists(p):
            return p
    return None

def _colblock_to_a1_range(spec: str) -> str:
    """
    'DEFG:4' -> 'D4:G4' ; 'NO:8' -> 'N8:O8' ; 'J:9' -> 'J9'
    Si ya viene 'A1:D1' u 'A1', lo regresa igual.
    """
    s = spec.strip().upper()
    if ":" not in s:
        return s
    left, right = s.split(":")
    # Caso 'A1:D1'
    if left and left[-1].isdigit():
        return s
    # Caso 'COLS:ROW'
    cols, row = left, right
    cols = "".join([c for c in cols if c.isalpha()])
    row = "".join([c for c in row if c.isdigit()])
    if not cols or not row:
        return s
    start = cols[0]
    end = cols[-1]
    if start == end:
        return f"{start}{row}"
    return f"{start}{row}:{end}{row}"

def _anchor_address(ws, addr: str) -> str:
    """
    Devuelve la coordenada superior-izquierda válida para escribir.
    Soporta celdas simples, rangos 'A1:D1' y nuestro formato 'DEFG:4'.
    """
    addr = _colblock_to_a1_range(addr)
    # Si es rango, usa su esquina superior-izquierda
    if ":" in addr:
        min_col, min_row, max_col, max_row = range_boundaries(addr)
        return f"{get_column_letter(min_col)}{min_row}"

    # Si es celda individual, pero pertenece a un merge -> ancla del merge
    cell = ws[addr]
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if addr in rng:
                min_col, min_row, *_ = range_boundaries(str(rng))
                return f"{get_column_letter(min_col)}{min_row}"
    return addr

# Bloques del template que queremos dejar en blanco si no hay dato en UI
CLEAR_BLOCKS = [
    "D4:G6",   # program/mould/material (entradas)
    "K4:O6",   # date/cavities/machine
    "J9:J9",   # Screw Ø [mm]
    "N8:O21",  # Key data
    "G11:I15", # Injection table
    "F18:F20", # Plasticizing
    "F23:H23","F24:H24", # Holding pressure
    "D27:H29", # Temperatures
    "E31:H33", # Opening
    "K31:N33", # Closing
    "D37:D37", # Clamping
]

def _clear_blocks(ws, blocks):
    """Limpia (pone None) sólo en la celda ancla si hay merges."""
    seen = set()
    for block in blocks:
        a1 = _colblock_to_a1_range(block)
        min_c, min_r, max_c, max_r = range_boundaries(a1)
        for col in range(min_c, max_c + 1):
            for row in range(min_r, max_r + 1):
                coord = f"{get_column_letter(col)}{row}"
                anchor = _anchor_address(ws, coord)
                if anchor not in seen:
                    ws[anchor].value = None
                    seen.add(anchor)

def _to_excel_value(excel_key: str, ui_key: str, raw_val):
    s = "" if raw_val is None else str(raw_val).strip()
    if s == "":
        return None  # celda en blanco si no diste valor

    # numérico si el campo es numérico
    if any(k in NUM_FIELDS for k in (excel_key, ui_key)):
        try:
            s2 = s.replace(",", "")
            v = float(s2)
            return int(v) if abs(v - int(v)) < 1e-12 else v
        except Exception:
            pass
    return _safe_excel(s)

# --- Fallback limpio con openpyxl ---
def _strip_links_and_drawings(wb):
    # Quitar vínculos externos para que Excel no “repare”
    try:
        wb._external_links = []
    except Exception:
        pass
    try:
        rels = getattr(wb, "_rels", None)
        if rels:
            for rid in list(rels):
                rel = rels[rid]
                if "externalLink" in getattr(rel, "reltype", ""):
                    del rels[rid]
    except Exception:
        pass

    # Quitar drawings/legacy drawings por hoja (imágenes/autoformas)
    for ws in wb.worksheets:
        try:
            ws._images = []
        except Exception:
            pass
        try:
            ws._charts = []
        except Exception:
            pass
        try:
            ws._drawing = None
        except Exception:
            pass
        try:
            if hasattr(ws, "_rels"):
                for rid in list(ws._rels):
                    if "drawing" in getattr(ws._rels[rid], "reltype", ""):
                        del ws._rels[rid]
        except Exception:
            pass
        try:
            ws.legacy_drawing = None
        except Exception:
            pass
        # Ocultar ceros en la vista
        try:
            ws.sheet_view.showZeros = False
        except Exception:
            pass

def _purge_all_formulas(wb):
    """Elimina TODA fórmula del libro (incluidas compartidas) para que Excel no 'repare'."""
    for ws in wb.worksheets:
        # romper metadatos de fórmulas compartidas
        try:
            ws.formula_attributes = {}
        except Exception:
            pass

        # recorrer rango usado y borrar fórmulas
        dim = ws.calculate_dimension()
        if not dim or ":" not in dim:
            continue
        min_c, min_r, max_c, max_r = range_boundaries(dim)
        for row in ws.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                try:
                    if getattr(cell, "data_type", None) == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
                        cell.value = None
                except Exception:
                    pass

    # algunos nombres definidos son fórmulas -> quitarlos
    try:
        keep = []
        for dn in list(wb.defined_names.definedName):
            if getattr(dn, "attr_text", "") and str(dn.attr_text).startswith("="):
                continue
            keep.append(dn)
        wb.defined_names.definedName = keep
    except Exception:
        pass

def _export_snapshot_to_template(snapshot: dict, out_path: str, template_path: str, sheet_name: str | None = None):
    from openpyxl import load_workbook
    keep_vba = template_path.lower().endswith((".xlsm", ".xlsb"))
    wb = load_workbook(template_path, data_only=False, keep_vba=keep_vba, keep_links=False)
    ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active

    # Limpiar bloques que mapeas
    _clear_blocks(ws, CLEAR_BLOCKS)

    # Escribir solo campos con valor
    for excel_key, spec in EXCEL_MAP.items():
        ui_key = ALIAS_EXCEL_TO_UI.get(excel_key, excel_key)
        val = snapshot.get(ui_key, "")
        target = _anchor_address(ws, spec)
        ws[target].value = _to_excel_value(excel_key, ui_key, val)

    # Eliminar externalLinks/drawings para que Excel NO "repairs"
    _strip_links_and_drawings(wb)

    # eliminar TODAS las fórmulas del libro (evita el “se han quitado fórmulas”)
    _purge_all_formulas(wb)

    wb.save(out_path)

def _a1_from_spec(spec: str) -> str:
    """Usa el parser "COLS:ROW" y toma la esquina superior-izquierda."""
    a1 = _colblock_to_a1_range(spec)
    return a1.split(":")[0] if ":" in a1 else a1

def _normalize_win_path(p: str) -> str:
    """Normaliza rutas en Windows: absolutas, con barras '\\' y sin espacios/puntos finales."""
    p = os.path.abspath(p).replace("/", "\\")
    return p.rstrip(" .")

def _export_with_excel_com(snapshot: dict, out_path: str, template_path: str, sheet_name: str | None = None):
    import tempfile, shutil, uuid
    import pywintypes  # viene con pywin32

    # --- normaliza y asegura carpeta destino ---
    out_path = _normalize_win_path(out_path)
    dest_dir = os.path.dirname(out_path)
    os.makedirs(dest_dir, exist_ok=True)

    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.AskToUpdateLinks = False

    wb = excel.Workbooks.Open(template_path, UpdateLinks=0, ReadOnly=False, CorruptLoad=0)
    try:
        excel.Calculation = -4135  # xlCalculationManual
        wb.CheckCompatibility = False

        ws = wb.Worksheets(sheet_name) if sheet_name else wb.Worksheets(1)

        for excel_key, spec in EXCEL_MAP.items():
            ui_key = ALIAS_EXCEL_TO_UI.get(excel_key, excel_key)
            raw = snapshot.get(ui_key, "")
            val = "" if raw is None or str(raw).strip() == "" else raw
            addr = _a1_from_spec(spec)
            ws.Range(addr).Value = val

        # --- guardado robusto: a %TEMP% y luego mover ---
        ext = ".xlsm" if out_path.lower().endswith((".xlsm", ".xlsb")) else ".xlsx"
        fmt = 52 if ext == ".xlsm" else 51
        tmp_path = os.path.join(tempfile.gettempdir(), f"~export_{uuid.uuid4().hex}{ext}")

        wb.SaveAs(Filename=tmp_path, FileFormat=fmt, ConflictResolution=2)

    except pywintypes.com_error as e:
        try:
            wb.Close(SaveChanges=False)
        except Exception:
            pass
        excel.Quit()
        raise

    wb.Close(SaveChanges=False)
    excel.Quit()

    try:
        if os.path.exists(out_path):
            os.remove(out_path)
    except PermissionError:
        # si está bloqueado, intenta sobreescritura atómica
        pass
    shutil.move(tmp_path, out_path)

# ---------- FS helpers ----------
def _safe_id(s: str) -> str:
    return str(s).replace("/", "_").replace("\\", "_").strip()

def _path_json(mold_id: str) -> str:
    return os.path.join(RECIPES_DIR, f"{_safe_id(mold_id)}.json")

def _versions_dir(mold_id: str) -> str:
    d = os.path.join(RECIPES_DIR, _safe_id(mold_id), "_versions")
    os.makedirs(d, exist_ok=True)
    return d

def _list_versions(mold_id: str):
    d = _versions_dir(mold_id)
    items = []
    rx = re.compile(r"^v(\d{3})\.json$")
    for fname in os.listdir(d):
        m = rx.match(fname)
        if m:
            items.append((m.group(0).replace(".json",""), os.path.join(d, fname)))
    items.sort(key=lambda t: t[0])
    return items

def _next_version_name(mold_id: str) -> str:
    items = _list_versions(mold_id)
    if not items: return "v001"
    last = items[-1][0]
    return f"v{int(last[1:])+1:03d}"

def _load_json(mold_id: str) -> dict:
    p = _path_json(mold_id)
    if os.path.exists(p):
        try:
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def _save_json(mold_id: str, data: dict):
    payload = dict(data)
    payload["_meta"] = {
        "schema": DATA_SCHEMA_VERSION,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    with open(_path_json(mold_id), "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

def _save_version_snapshot(mold_id: str, data: dict, usuario: str, motivo: str, diffs_text: str) -> str:
    ver = _next_version_name(mold_id)
    snap = dict(data)
    snap["_meta"] = {
        "schema": DATA_SCHEMA_VERSION,
        "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "usuario": usuario,
        "motivo": motivo,
        "version": ver,
        "diffs": diffs_text,
    }
    out = os.path.join(_versions_dir(mold_id), f"{ver}.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(snap, f, ensure_ascii=False, indent=2)
    return ver

def _load_version_snapshot(mold_id: str, version: str) -> dict:
    path = os.path.join(_versions_dir(mold_id), f"{version}.json")
    if not os.path.exists(path): return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def _cast_numeric(val: str) -> str:
    if val is None: return ""
    s = str(val).strip().replace(",", ".")
    if s == "": return ""
    try:
        num = float(s)
        return str(int(num)) if abs(num - int(num)) < 1e-9 else f"{num:.3f}".rstrip("0").rstrip(".")
    except Exception:
        return str(val).strip()

def _ascii(s: str) -> str:
    """Evita errores latin-1 al exportar PDFs en algunos backends."""
    if s is None: return ""
    return (str(s).replace("\u2014","-").encode("latin-1","replace").decode("latin-1"))

# ---------- Sanitizers para export ----------
try:
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
except Exception:
    ILLEGAL_CHARACTERS_RE = None

def _safe_excel(val: str) -> str:
    s = "" if val is None else str(val)
    if ILLEGAL_CHARACTERS_RE:
        s = ILLEGAL_CHARACTERS_RE.sub("", s)
    return s

def _safe_pdf(val: str) -> str:
    # Latin-1 friendly + sin saltos raros
    return _ascii(val).replace("\r", " ").replace("\n", " ")
